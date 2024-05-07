import os.path
from .pytmysql import PyMySQL
import xlrd,xlsxwriter
from .bk_179 import optstr
from file_ls import ls
from tqdm import tqdm

def excel_into_mysql(file:str,host:str, password:str, database:str,bm:str,th_a=1,stat:int=0,port:int=3306, user:str='root',auto_add_key:bool=True):
    """
    :param file: 文件路径
    :param host: mysql地址
    :param password: mysql密码
    :param database: mysql数据库名称
    :param bm: 数据库表名称
    :param th_a: 插入数据表表头，可传入excel对应行或列表
    :param stat: 插入数据库数据行，有默认规则，可自定义
    :param port: mysql端口号
    :param user: mysql用户名
    :param auto_add_key: 是否自动添加表头字段
    :return:
    """
    data = xlrd.open_workbook(file)
    table = data.sheets()[0]  #通过工作表索引顺序获取
    nrows = table.nrows #有效行数
    ncols = table.ncols #有效列数
    data=[]
    for r in range(nrows):
        l=[]
        for c in range(ncols):
            l.append(table.cell_value(r, c))
        data.append(l)
    mysql = PyMySQL(host=host,port=3306, user='root', password=password, database=database,auto_add_key=auto_add_key)
    th=[]
    if type(th_a) is int:
        th=data[th_a-1]
        stat=th_a
    elif type(th) is list:
        th = th_a
    for v in data[stat:]:
        mysql.insert_into(bm,th,v)
    mysql.close()

def gen_excel(data:list,save_path,header=None,is_optstr=True):
    """
    :param data:列表数据
    :param save_path:保存excel路径
    :param header:表头
    """
    if header:
        data = list(data)
        data.insert(0,header)
    workbook = xlsxwriter.Workbook(save_path if save_path.endswith('.xlsx') else f'{save_path}.xlsx')
    sheet = workbook.add_worksheet()
    for row_num, row_data in enumerate(tqdm(data,desc=f'{os.path.basename(save_path)}')):
        for col_num, col_value in enumerate(row_data):
            if is_optstr:
                sheet.write(row_num, col_num, optstr(col_value if col_value is None else str(col_value)))
            else:
                sheet.write(row_num, col_num, col_value)
    workbook.close()

#保留有效的小数位
def prvadepl(num):
    if type(num)==float or type(num)==numpy.float64:
        num = str(num).rstrip('0').rstrip('.')
        if '.' in num:
            num = float(num)
        else:
            num = int(num)
    return num

#读取execl数据
import os.path
from .pytmysql import PyMySQL
import xlrd,xlsxwriter,zipfile
from .bk_179 import optstr
from file_ls import ls
from tqdm import tqdm

def excel_into_mysql(file:str,host:str, password:str, database:str,bm:str,th_a=1,stat:int=0,port:int=3306, user:str='root',auto_add_key:bool=True):
    """
    :param file: 文件路径
    :param host: mysql地址
    :param password: mysql密码
    :param database: mysql数据库名称
    :param bm: 数据库表名称
    :param th_a: 插入数据表表头，可传入excel对应行或列表
    :param stat: 插入数据库数据行，有默认规则，可自定义
    :param port: mysql端口号
    :param user: mysql用户名
    :param auto_add_key: 是否自动添加表头字段
    :return:
    """
    data = xlrd.open_workbook(file)
    table = data.sheets()[0]  #通过工作表索引顺序获取
    nrows = table.nrows #有效行数
    ncols = table.ncols #有效列数
    data=[]
    for r in range(nrows):
        l=[]
        for c in range(ncols):
            l.append(table.cell_value(r, c))
        data.append(l)
    mysql = PyMySQL(host=host,port=3306, user='root', password=password, database=database,auto_add_key=auto_add_key)
    th=[]
    if type(th_a) is int:
        th=data[th_a-1]
        stat=th_a
    elif type(th) is list:
        th = th_a
    for v in data[stat:]:
        mysql.insert_into(bm,th,v)
    mysql.close()

def gen_excel(data:list,save_path,header=None,is_optstr=True):
    """
    :param data:列表数据
    :param save_path:保存excel路径
    :param header:表头
    """
    if header:
        data = list(data)
        data.insert(0,header)
    workbook = xlsxwriter.Workbook(save_path if save_path.endswith('.xlsx') else f'{save_path}.xlsx')
    sheet = workbook.add_worksheet()
    for row_num, row_data in enumerate(tqdm(data,desc=f'{os.path.basename(save_path)}')):
        for col_num, col_value in enumerate(row_data):
            if is_optstr:
                sheet.write(row_num, col_num, optstr(col_value if col_value is None else str(col_value)))
            else:
                sheet.write(row_num, col_num, col_value)
    workbook.close()

#保留有效的小数位
def prvadepl(num):
    if type(num)==float or type(num)==numpy.float64:
        num = str(num).rstrip('0').rstrip('.')
        if '.' in num:
            num = float(num)
        else:
            num = int(num)
    return num

#读取execl数据
def read_excel(file_path,sheet_index=0):
    """
    :param file_path:文件路径或路径下第一个文件
    :param sheet_index:工作表索引默认第一个
    return:表格数据
    """
    if not file_path.endswith('.xlsx'):
        file_paths=ls(file_path)
        for path in file_paths:
            if path.endswith('.xlsx'):
                file_path=path
                break
    from openpyxl import load_workbook
    try:
        workbook = load_workbook(filename=file_path,read_only=True)
        worksheet = workbook.worksheets[sheet_index]
        datas = []
        for row in worksheet.iter_rows():
            datas.append([prvadepl(cell.value) if isinstance(cell.value, float) else cell.value for cell in row])
        return datas
    except zipfile.BadZipFile:
        from txdpy import webptablesl
        with open(file_path, 'r', encoding='utf-8') as f:
            return webptablesl(f.read(), '//table')