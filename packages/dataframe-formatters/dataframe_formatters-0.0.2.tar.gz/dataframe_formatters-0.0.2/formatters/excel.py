import copy
from pathlib import Path
from string import ascii_uppercase
from typing import (
    Sequence,
    Literal as L,
)

import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Border,
    Font,
    NamedStyle,
    Side,
)


def prettify_workbook(
    *,
    infile: Path,
    outfile: Path,
    ws_names: Sequence[str],
) -> None:
    """Format excel."""
    wb = openpyxl.load_workbook(infile.as_posix())

    styles = _add_style(wb)
    worksheets = [wb.active]
    for ws, name in zip(worksheets, ws_names):
        ws = _prettify_excel_ws(ws, name, styles)

    wb.save(outfile.as_posix())


def _add_style(wb: openpyxl.Workbook) -> dict[L["body", "header"], NamedStyle]:
    def header() -> NamedStyle:
        header = NamedStyle(name="header")
        wb.add_named_style(header)
        header_bd = Side(style="medium")
        header.border = Border(
            left=header_bd,
            top=header_bd,
            right=header_bd,
            bottom=header_bd,
        )
        header.font = Font(name="Verdana", sz=11, bold=True)
        return header

    def body() -> NamedStyle:
        body = NamedStyle(name="body")
        wb.add_named_style(body)
        bd = Side(style="thin")
        body.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        return body

    return {"header": header(), "body": body()}


def _prettify_excel_ws(
    ws: Worksheet, name: str, styles: dict[str, NamedStyle]
) -> Worksheet:
    ws = copy.copy(ws)
    ws.title = name

    def set_width(ws: Worksheet):
        df = pd.DataFrame(ws.values).astype(str)
        df.columns = df.iloc[0]

        for col_letter, col in zip(ascii_uppercase, df.columns):
            avg_width = df[col].apply(lambda x: len(x)).mean()
            ws.column_dimensions[col_letter].width = max(avg_width * 1.5, len(str(col)))

    def set_style(ws: Worksheet):
        for cell in ws["1:1"]:
            cell.style = styles["header"]

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.style = styles["body"]

    set_width(ws)
    set_style(ws)

    return ws
