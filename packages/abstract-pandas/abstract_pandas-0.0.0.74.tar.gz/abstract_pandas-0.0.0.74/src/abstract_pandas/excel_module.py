import os
import pandas as pd
from abstract_utilities import *
from .general_functions import split_and_clean_lines,make_type,convert_column,return_float_or_int,safe_get,get_number
import geopandas as gpd
from openpyxl import load_workbook, utils
from difflib import get_close_matches
from werkzeug.utils import secure_filename
from werkzeug.datastructures import FileStorage
import logging
# Configure logging at the start of your script or application
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
def source_ext(typ=None):
    source_js = {'.parquet':'pyarrow','.txt':'python','.xlsx':'openpyxl','.xls':'openpyxl','.xlsb':'pyxlsb','.ods':'odf','.geojson':'GeoJSON'}
    if typ:
        source_js = source_js.get(typ)
    return source_js
def isDataFrame(obj):
    return isinstance(obj, pd.DataFrame)
def create_dataframe(new_data=None,columns=None):
    if isDataFrame(new_data):
        return new_data
    new_data = new_data or {}
    if isinstance(new_data,dict):
        new_data=[new_data]
        if columns == None:
            columns=[]
            for datas in new_data:
                if isinstance(datas,dict):
                    columns=list(set(columns+list(datas.keys())))
        if columns ==False:
            columns=None
    if isinstance(new_data,list):
        return pd.DataFrame(new_data,columns=columns)
def read_excel_range(df, start_row, num_rows):
    # Skip rows up to start_row, not including the header if it's the first row.
    # Adjust start_row by 1 if your Excel file has headers and you want to include them.
    # start_row is 0-indexed in Python, but 1-indexed in Excel, so adjust accordingly.
    skip = start_row - 1 if start_row > 0 else None
    
    # Read the specified range of rows
    df = get_df(df, skiprows=skip, nrows=num_rows, header=None if skip is None else 0)
    
    return df
def count_rows_in_excel(data_source, sheet_name=None):
    """
    Count rows in a given Excel sheet or pandas DataFrame.

    Parameters:
    - data_source (str or pd.DataFrame): The file path of the Excel file or a pandas DataFrame.
    - sheet_name (str, optional): The name of the sheet to count rows in. Used only if data_source is a file path.

    Returns:
    - int: Number of rows in the Excel sheet or DataFrame.
    """
    if isinstance(data_source, pd.DataFrame):
        # If data_source is a DataFrame, simply return the number of rows
        return len(data_source)
    elif isinstance(data_source, str):
        # Assume data_source is a file path to an Excel file
        try:
            wb = load_workbook(filename=data_source, read_only=True)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb[wb.sheetnames[0]]

            row_count = ws.max_row
            if row_count == 1 and ws.max_column == 1:
                if ws.cell(row=1, column=1).value is None:
                    row_count = 0
            
            wb.close()
            return row_count
        except utils.exceptions.InvalidFileException:
            print("Failed to open the file. It may be corrupted or the path is incorrect.")
            return None
        except Exception as e:
            print(f"An error occurred: {e}")
            return None
    else:
        print("Invalid input. Please provide a valid file path or pandas DataFrame.")
        return None

def get_df(source=None, nrows=None, skiprows=None, condition=None, indices=None):
    """
    Load a DataFrame from various sources with optional filtering.

    Parameters:
    - source (str, pd.DataFrame, gpd.GeoDataFrame, dict, list, FileStorage): Source of the data.
    - nrows (int, optional): Number of rows to load.
    - header (int, list of int, 'infer', optional): Row(s) to use as the header.
    - skiprows (list-like, int, optional): Rows to skip at the beginning.
    - condition (pd.Series, optional): Condition for filtering rows.
    - indices (list of int, optional): Indices of rows to select.

    Returns:
    - pd.DataFrame or gpd.GeoDataFrame: Loaded and optionally filtered data.
    """
    if isinstance(source, (pd.DataFrame, gpd.GeoDataFrame)):
        logging.info("Data is already loaded as a DataFrame/GeoDataFrame.")
        return filter_df(source, nrows=nrows, condition=condition, indices=indices)

    if source is None:
        logging.error("No source provided for loading data.")
        return None

    if isinstance(source, str) and os.path.isfile(source):
        file_ext = os.path.splitext(source)[-1].lower()
        try:
            logging.info(f"Loading data from file with extension: {file_ext}")
            if file_ext in ['.csv', '.tsv', '.txt']:
                sep = {'csv': ',', 'tsv': '\t'}.get(file_ext.strip('.'), None)
                df = pd.read_csv(source,  skiprows=skiprows, sep=sep, nrows=nrows)
            elif file_ext in ['.ods', '.xlsx', '.xls', '.xlsb']:
                engine = {'ods': 'odf', 'xlsx': 'openpyxl', 'xls': 'xlrd', 'xlsb': 'pyxlsb'}.get(file_ext.strip('.'))
                df = pd.read_excel(source,  skiprows=skiprows, engine=engine, nrows=nrows)
            elif file_ext == '.json':
                df = pd.read_json(source, nrows=nrows)
            elif file_ext == '.parquet':
                df = pd.read_parquet(source, nrows=nrows)
            elif file_ext in ['.shp', '.cpg', '.dbf', '.shx','.geojson']:
                df = gpd.read_file(source,driver=source_ext(file_ext))
            elif file_ext in ['.prj']:
                df = read_from_file(source)
                return df
            else:
                try:
                    df = read_from_file(source)
                except:                
                    raise logging.info(f"Unsupported file extension: {file_ext}")
            if not isinstance(df, (dict, list,FileStorage)):
                return filter_df(df, nrows=nrows, condition=condition, indices=indices)
            source = df
        except Exception as e:
            logging.error(f"Failed to read file: {e}")
            return None

    if isinstance(source, FileStorage):
        try:
            logging.info(f"Reading from FileStorage: {secure_filename(source.filename)}")
            df = pd.read_excel(source.stream, nrows=nrows)
            return filter_df(df, nrows=nrows, condition=condition, indices=indices)
        except Exception as e:
            logging.error(f"Failed to read from FileStorage: {e}")
            return None

    if isinstance(source, (dict, list)):
        logging.info("Creating DataFrame from in-memory data structure.")
        df = pd.DataFrame(source)
        return filter_df(df, nrows=nrows, condition=condition, indices=indices)

    logging.error("Invalid data source type provided.")
    return None

def filter_df(df, nrows=None, condition=None, indices=None):
    """
    Apply filtering to a DataFrame based on specified criteria.

    Parameters:
    - df (DataFrame): The DataFrame to filter.
    - nrows (int, optional): Number of rows to return from the start.
    - condition (pd.Series, optional): Boolean series for row filtering.
    - indices (list of int, optional): Row indices to select.

    Returns:
    - DataFrame: Filtered DataFrame.
    """
    if nrows is not None:
        df = df.head(nrows)
    if condition is not None:
        df = df[condition]
    if indices is not None:
        df = df.iloc[indices]
    return df


def update_or_append_data(df_new_data=None, df_existing_data=None, search_column=None, search_value=None, clear_duplicates=False):
    df_new = get_df(df_new_data)
    df_existing = get_df(df_existing_data)
    
    if df_existing.empty:
        return df_new

    # Ensure new data columns exist in the existing dataframe, add if not
    for col in df_new.columns:
        if col not in df_existing.columns:
            df_existing[col] = pd.NA  # Use pd.NA for missing data

    if search_column and search_value:
        if isinstance(search_column, list) and isinstance(search_value, list):
            mask = pd.Series(True, index=df_existing.index)
            for col, val in zip(search_column, search_value):
                mask &= (df_existing[col] == val)
        else:
            mask = (df_existing[search_column] == search_value)

        if mask.any():
            # Update existing rows based on mask
            for col in df_new.columns:
                df_existing.loc[mask, col] = df_new.loc[df_new.index[0], col]
            print(f"Updated rows where {search_column} matches {search_value}.")
        else:
            # Append new data if no matching row is found
            df_existing = pd.concat([df_existing, df_new], ignore_index=True)
            print(f"Appended new data as no existing match found for {search_value}.")

        # Handle duplicates update
        if not clear_duplicates:
            first_indices = df_existing.drop_duplicates(subset=search_column, keep='first').index
            update_mask = ~df_existing.index.isin(first_indices)
            for col in df_existing.columns:
                df_existing.loc[update_mask, col] = df_existing.loc[df_existing[df_existing[search_column] == df_existing.loc[update_mask, search_column]].index[0], col]

            print("Updated duplicates to match the first occurrence.")
    else:
        df_existing = pd.concat([df_existing, df_new], ignore_index=True)
        print("Appended new data as no search criteria provided.")

    if clear_duplicates:
        df_existing = df_existing.drop_duplicates()
        print("Duplicates removed after update/append.")

    return df_existing



def get_cell_value(df, column_header, row_index=1):
    """
    Retrieves the value from a specified cell in the GeoDataFrame.
    
    :param df: GeoDataFrame or filepath to the shapefile.
    :param column_header: The header of the column from which to retrieve the value.
    :param row_index: The index of the row from which to retrieve the value.
    :return: The value located at the specified column and row.
    """
    df = get_df(df)
    # Check if the column header is in the GeoDataFrame
    if column_header not in df.columns:
        raise ValueError(f"The column header '{column_header}' does not exist in the GeoDataFrame.")
    
    # Check if the row index is within the bounds of the GeoDataFrame
    if not (0 <= int(row_index) < len(df)):
        raise ValueError(f"The row index {row_index} is out of bounds for the GeoDataFrame.")
    
    # Retrieve and return the value from the specified cell
    return df.iloc[row_index][column_header]
def convert_value(value, column_dtype):
    """ Convert the value to the column data type if possible, handling string representations of numbers. """
    if pd.api.types.is_numeric_dtype(column_dtype):
        try:
            return float(value)
        except ValueError:
            return value  # Return the original value if conversion fails
    return value

def search_df_for_values(df, column_name, query_list, type_dependent=False):
    """
    Search DataFrame column for rows matching any items in query_list with optional type-dependent matching.

    Parameters:
    - df (pd.DataFrame): The DataFrame to search.
    - column_name (str): The name of the column to search.
    - query_list (list or single value): A list of values or a single value to search for in the column.
    - type_dependent (bool): Whether to enforce type matching.

    Returns:
    - pd.DataFrame: A DataFrame of rows where the column values match any item in the query_list.
    """
    df = pd.DataFrame(df)  # Ensure it is a DataFrame
    query_list = make_list(query_list)

    if type_dependent:
        # Enforcing exact type matching
        mask = df[column_name].apply(lambda x: any([x == item and type(x) == type(item) for item in query_list]))
    else:
        # Attempt to convert query values to the column data type for accurate comparison
        column_dtype = df[column_name].dtype
        converted_query_list = [convert_value(item, column_dtype) for item in query_list]
        mask = df[column_name].isin(converted_query_list)

    return df[mask]
def add_or_update_headers(df, column_name, default_value=None):
    """
    Add a new column to a DataFrame with a default value if it does not already exist.

    Parameters:
    df (DataFrame): The DataFrame to modify.
    column_name (str): The name of the column to add.
    default_value (Any, optional): The default value to assign to the new column. Defaults to None.

    Returns:
    DataFrame: The modified DataFrame with the new column added if it didn't exist.
    """
    if column_name not in df.columns:
        df[column_name] = default_value
    else:
        print(f"Column '{column_name}' already exists in the DataFrame. No changes made.")

    return df
def search_df_with_condition(df, column_name, condition_func):
    """
    Search DataFrame column to find rows where condition_func returns True.

    Parameters:
    - df (pd.DataFrame): The DataFrame to search.
    - column_name (str): The column to apply the condition on.
    - condition_func (function): A function that takes a single value and returns True or False.

    Returns:
    - pd.DataFrame: A DataFrame of rows where the column values satisfy the condition_func.
    """
    df=get_df(df)
    # Applying the condition function vectorized
    mask = df[column_name].apply(condition_func)
    return df[mask]
def query_dataframe(df, query_string):
    """
    Use DataFrame.query() to filter rows based on a query string.

    Parameters:
    - df (pd.DataFrame): The DataFrame to query.
    - query_string (str): The query string to evaluate.

    Returns:
    - pd.DataFrame: The filtered DataFrame.
    """
    return df.query(query_string)


def filter_and_deduplicate_df(df, filter_columns, filter_values, dedup_columns=None):
    """
    Filters a DataFrame based on specified values in given columns and removes duplicates.

    Parameters:
    - df (pd.DataFrame): The DataFrame to filter and deduplicate.
    - filter_columns (list of str): Column names to apply the filters on.
    - filter_values (list of list): Lists of values to include for each column in filter_columns.
    - dedup_columns (list of str, optional): Columns to consider for dropping duplicates. If not specified,
      duplicates will be dropped based on all columns.

    Returns:
    - pd.DataFrame: The filtered and deduplicated DataFrame.
    """
    # Ensure the input integrity
    assert len(filter_columns) == len(filter_values), "Each filter column must correspond to a list of filter values."

    # Apply filters based on the columns and corresponding values
    mask = pd.Series([True] * len(df))
    for col, vals in zip(filter_columns, filter_values):
        mask &= df[col].isin(vals)

    filtered_df = df[mask]

    # Drop duplicates based on specified columns
    if dedup_columns:
        deduplicated_df = filtered_df.drop_duplicates(subset=dedup_columns)
    else:
        deduplicated_df = filtered_df.drop_duplicates()

    return deduplicated_df

def count_rows_columns(df):
    """
    Counts the number of rows and columns in a pandas DataFrame.

    Parameters:
    - df (DataFrame): The pandas DataFrame whose dimensions will be counted.

    Returns:
    - tuple: A tuple containing two elements, the number of rows and the number of columns in the DataFrame.
    """
    rows, columns = df.shape  # df.shape returns a tuple (number of rows, number of columns)
    return rows, columns
def get_min_max_from_query(query):
    query = [return_float_or_int(obj) for obj in make_list(query) if is_number(obj)]
    query.sort()
    minimum = get_number(safe_get(query,0))
    maximum = get_number(safe_get(query,-1))
    minimum = minimum if len(query) > 0 and minimum is not None else 0
    maximum = maximum if len(query) > 1 and maximum is not None else 900
    return minimum,maximum
def is_inverse(series,inverse=False):
    if inverse:
        series = ~series
    return series
def get_geo_location_master():
    return "/home/catalystRepository/distance_repository/filters/geo_location/zip_codes/sacramento_to_zipcodes_geo_locations.xlsx"
def get_geo_location_master_headers():
    return """COUNTY	COUNTY_closest_latitude	COUNTY_closest_longitude	ZIP	ZIP_closest_latitude	ZIP_closest_longitude	ZIP_COUNTY_closest_distance(mi)	ZIP_furthest_latitude	ZIP_furthest_longitude	ZIP_COUNTY_furthest_distance(mi)""".split('\t')
async def distance_within_range(distance,geo_location_refference=None):
    logging.info(f"distance = {distance}\n\geo_location_refference={geo_location_refference}")
    geo_location_refference = geo_location_refference or get_geo_location_master()
    # Ensure distance is a float
    distance = float(get_number(safe_get(distance, 0)))
    
    # Load geo location data
    df_geo_location = get_df(geo_location_refference)
    
    # Convert ZIP codes to integers, assuming all entries are valid zip codes
    df_geo_location['ZIP'] = df_geo_location['ZIP'].astype(str)
    
    # Convert distances to floats
    df_geo_location['ZIP_COUNTY_furthest_distance(mi)'] = df_geo_location['ZIP_COUNTY_furthest_distance(mi)'].astype(float)
    df_geo_location['ZIP_COUNTY_closest_distance(mi)'] = df_geo_location['ZIP_COUNTY_closest_distance(mi)'].astype(float)
    
    # Create a mask where the distance is within the specified range
    mask = (distance >= df_geo_location['ZIP_COUNTY_closest_distance(mi)']) | (distance >= df_geo_location['ZIP_COUNTY_furthest_distance(mi)'])
    logging.info(f"mask = {mask}")

    # Filter zip codes based on the mask
    good_zips = df_geo_location.loc[mask, 'ZIP'].tolist()
    good_zips=['00000'[:-len(zips)]+str(zips) for zips in good_zips]
    return good_zips

async def get_range(df, column,query,inverseOption=False,caseOption=False,substringOption=False):
    """
    Filters the DataFrame based on numeric ranges specified in query.
    Optionally inverts the filter to exclude the specified range.
    
    :param df: DataFrame or path to DataFrame
    :param column: Column name to apply the range filter on
    :param query: List or tuple containing the minimum and maximum values as strings
    :param invert: If True, the range is inverted (excludes the range specified)
    :return: Filtered DataFrame
    """
    if not isinstance(query,list):
        query = split_and_clean_lines(query)
    logging.info(f"get_range processing")
    temp_df = df.copy()
    logging.info(f"temp_df created with type {type(temp_df)}")
    minimum,maximum = get_min_max_from_query(query)
    logging.info(f"minimum,maximum : {minimum},{maximum}")
    temp_df[column] = pd.to_numeric(temp_df[column], errors='coerce')  # Coerce errors in case of non-numeric data
    # Create a condition for values within the specified range
    condition = temp_df[column].between(minimum,maximum , inclusive='both')
    # Filter the original DataFrame using the condition from the temporary DataFrame
    series = is_inverse(condition,inverse=inverseOption)
    return df[series]


async def filter_dataframe(df, column,query,caseOption=False,substringOption=False,inverseOption=False):
    """
    Filters a DataFrame based on a comparison between a column's values and a list.

    Parameters:
    - df: Pandas DataFrame.
    - column: Column name as a string where the comparisons are to be made.
    - compare_list: List of strings to compare against the DataFrame's column.
    - substring_match: Boolean, if True performs substring matching, otherwise exact matching.
    - case_sensitive: Boolean, if False converts both column and list_obj to lowercase.

    Returns:
    - filtered_df: DataFrame containing only the rows that meet the condition.
    """
    if not isinstance(query,list):
        query = split_and_clean_lines(query)
    compare_list=make_list(query)
    logging.info(f"filter_dataframe processing")
    if not caseOption:
        compare_list = [str(item).lower() for item in compare_list]
        df[column] = df[column].astype(str).str.lower()

    if substringOption:
        condition = df[column].apply(lambda x: any(str(item) in str(x) for item in compare_list))
    else:
        condition = df[column].apply(lambda x: any(str(item) == str(x) for item in compare_list))
    series = is_inverse(condition,inverse=inverseOption)
    return df[series]
async def filter_excel(df,column,query,range_option=False,caseOption=False,substringOption=False,inverseOption=False,distanceOption=False,geo_location_refference=None):
    logging.info(f"recieved excel: parsing for column={column},range_option={range_option},caseOption={caseOption},substringOption={substringOption},inverseOption={inverseOption},distanceOption={distanceOption},geo_location_refference={geo_location_refference}")
    df = get_df(df)
    logging.info(f"query came as {query}")
    if not isinstance(query,list):
        query = split_and_clean_lines(query)
    query=make_list(query)
    logging.info(f"query started as  as {query}")
    if distanceOption:
        logging.info(f"distanceOption selected")
        column=column or 'ZIP'
        query = await distance_within_range(distance=query,geo_location_refference=geo_location_refference)
        logging.info(f"query is now  as {query}")
    if range_option:
        logging.info(f"rangeOption selected")
        df = await get_range(df,column,query,inverseOption=inverseOption,caseOption=caseOption,substringOption=substringOption)
    else:
        df = await filter_dataframe(df,column,query,caseOption=caseOption,substringOption=substringOption,inverseOption=inverseOption)
    return df
def update_excel(df,header,query):
    df = get_df(df)
    query = make_type(query,[str])
    df = convert_column(df,header,[str])
    return df,query
def merge_dataframes(dataframes):
    merged_df = pd.concat(dataframes, ignore_index=True)
    return merged_df
def excel_to_dict(df):
    # Read the Excel file
    df = get_df(df)
    # Convert each row to a dictionary with column headers as keys
    rows_as_dicts = df.to_dict(orient='records')
    return rows_as_dicts
def get_headers(df):
    df = get_df(df)
    column_names = df.columns.tolist()
    return column_names
def get_row_as_list(df,index=0):
    df=get_df(df)
    if get_row_number(df)>index:
        return df.loc[index].astype(str).tolist()
def get_row_number(df):
    df=get_df(df)
    return len(df)
def get_int_from_column(obj,column):
    headers = get_headers(obj)
    return get_itter(headers,column)
def get_column_from_int(obj,i):
    headers = get_headers(obj)
    column = safe_itter_get(headers,i)
    # Return the column headers
    return column
def find_row_with_matching_cell(excel_datas,search_column='',search_value=''):
    matching_row = [excel_data for excel_data in excel_datas if isinstance(excel_data,dict) and excel_data.get(search_column) == search_value]
    if matching_row and isinstance(matching_row,list) and len(matching_row)>0:
        return matching_row[0]
    return {}
def get_itter(list_obj,target):
    for i,string in enumerate(list_obj):
        if string == target:
            return i
    return None
def safe_itter_get(list_obj,i):
    if len(list_obj)>i:
        return list_obj[i]
def get_expected_headers(df,*expected_headers):
    df = get_df(df)
    if isinstance(expected_headers,tuple or set):
        expected_headers=list(expected_headers)
    else:
        expected_headers=make_list(expected_headers)
    expected_headers = {expected_header:"" for expected_header in expected_headers}
    return get_closest_headers(df,expected_headers)
def get_closest_headers(df,expected_headers={}):
    actual_headers = get_headers(df)  # Extract the actual headers from the DataFrame
    # Mapping actual headers to expected headers based on closest match
    for expected_header in expected_headers:
        # Using get_close_matches to find the closest match; returns a list
        close_matches = get_close_matches(expected_header, actual_headers, n=1, cutoff=0.6)
        if close_matches:
            expected_headers[expected_header] = close_matches[0]
        else:
            # If no close matches found, leave as empty string which signifies no match found
            expected_headers[expected_header] = ""
    return expected_headers
def get_first_for_each(df,headers,queries,new_file_path=None):
    df=get_df(df)
    headers = get_expected_headers(df,headers).values()
    # Filter the DataFrame to only include rows with ZIP codes that are in the 'zips' list
    df=filter_and_deduplicate_df(df, headers, queries, dedup_columns=None)
    # Save the filtered and deduplicated DataFrame to a new Excel file
    return df

def get_row(target_value,column_name=None,index_value=None,df=None):
    # Use the `.str.contains` method directly in the indexing to filter rows
    df = get_df(df)
    column_names = make_list(column_name or df.columns.tolist())
    for i,column in enumerate(column_names):
        search_column(target_value,column_name=None,column_index=0,df=None,exact_match=True)
        value = df[column][index_value]
        if value.lower() == target_value.lower():
            return df,i
    return None,None
def search_column(target_value,column_name=None,column_index=0,df=None,exact_match=True):
    df = get_df(df)
    headers = df.columns.tolist()
    if not column in headers and column_index >len(headers):
        return
    column_values = df[column].tolist()
    if target_value in column_values:
        return get_itter(column_values=column_values,target_value=target_value,exact_match=exact_match)
    if not exact_match:
        if [val for val in column_values if str(target_value) in str(val)]:
           return get_itter(column_values=column_values,target_value=target_value,exact_match=exact_match)
    column_names = make_list(column_name or headers)
    return column_names
def add_data_point(df, data_point,header=None):
    """
    Adds a data point and its header to the next available column in the Excel file.

    Parameters:
        file_path (str): The path to the Excel file.
        header (str): The header for the data point.
        data_point (str or numeric): The data point to add.
    
    Returns:
        None: The function updates the Excel file directly.
    """
    # Ensure the file exists or create a new DataFrame if it doesn't
    df = get_df(df)
    # Check if the header already exists
    if header in df.columns:
        # Find first empty row in the existing column
        empty_row = df[header].isna().idxmax() if not df[header].dropna().empty else 0
        df.at[empty_row, header] = data_point
    else:
        # Add new column with the header and place the data point at the first row
        df[header] = pd.Series([data_point] + [None] * (len(df) - 1))
    
    # Save the updated DataFrame back to the Excel file
    return df
