import openpyxl
import os


def separate(voucher_file_path, voucher_code_column_name, account_code_column_name, sheet_name=None, result_folder_path=None):

    if sheet_name is None:
        sheet_name = ''

    if result_folder_path is None:
        result_folder_path = ''

    if os.path.isfile(voucher_file_path):
        basename = os.path.basename(voucher_file_path)
        name, extension = os.path.splitext(basename)

        if not extension.lower() == '.xlsx':
            raise ValueError('파일은 .xlsx 확장자이어야 합니다. ', basename)

    else:
        raise FileNotFoundError('파일을 찾을 수 없습니다. ', voucher_file_path)

    print('다음 경로의 파일을 읽습니다. ', voucher_file_path)

    wb = openpyxl.load_workbook(voucher_file_path)

    print('파일을 성공적으로 읽었습니다.')

    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    print('다음 이름의 시트를 읽습니다. ', sheet_name)

    ws = wb[sheet_name]

    print('시트를 성공적으로 읽었습니다.')

    start_row = 0

    voucher_code_column_index = 0
    account_code_column_index = 0

    for row in ws.iter_rows():

        found_voucher_code_column_name = False
        found_account_code_column_name = False

        for cell in row:
            if str(cell.value).strip().replace(' ', '').replace(r'\n', '') == voucher_code_column_name:
                found_voucher_code_column_name = True
                voucher_code_column_index = cell.column
            if str(cell.value).strip().replace(' ', '').replace(r'\n', '') == account_code_column_name:
                found_account_code_column_name = True
                account_code_column_index = cell.column

            if found_account_code_column_name and found_voucher_code_column_name:
                start_row = cell.row
                break

    if start_row == 0:
        raise ValueError(f'시트 "{sheet_name}" 에서 전표번호 열 이름 "{voucher_code_column_name}" 과 계정코드 열 이름 "{account_code_column_name}" 이 모두 존재하는 행을 찾을 수 없습니다')

    account_d = {}
    voucher_d = {}

    max_row = ws.max_row
    max_column = ws.max_column

    for index_row in range(start_row + 1, max_row):
        voucher_code = str(ws.cell(index_row, voucher_code_column_index).value)
        account_code = str(ws.cell(index_row, account_code_column_index).value)

        if account_code not in account_d:
            account_d[account_code] = [voucher_code]

        else:
            if voucher_code not in account_d[account_code]:
                account_d[account_code].append(voucher_code)

        if voucher_code not in voucher_d:
            voucher_d[voucher_code] = [index_row]
        else:
            voucher_d[voucher_code].append(index_row)

    if not result_folder_path:
        result_folder_path = os.path.dirname(voucher_file_path)

    account_code_count = len(list(account_d.keys()))

    index_account = 0

    for account_code in account_d:
        index_account += 1
        print(f'{account_code}.xlsx', '파일을 생성 중입니다.', index_account, '/', account_code_count)
        wb_new = openpyxl.Workbook()
        ws_new = wb_new[wb_new.sheetnames[0]]
        index_row = 1
        for j in range(0, max_column):
            index_column = j + 1
            ws_new.cell(index_row, index_column).value = ws.cell(start_row, index_column).value
        for voucher_code in account_d[account_code]:
            for index_row_original in voucher_d[voucher_code]:
                index_row += 1
                for j in range(0, max_column):
                    index_column = j + 1
                    ws_new.cell(index_row, index_column).value = ws.cell(index_row_original, index_column).value

        result_file_path = os.path.join(result_folder_path, f'{account_code}.xlsx')
        wb_new.save(result_file_path)

    return 0


if __name__ == '__main__':
    separate(
        voucher_file_path=r"23.12 분개장(23.01~12).xlsx",
        voucher_code_column_name='전표승인번호',
        account_code_column_name='계정코드'
    )
