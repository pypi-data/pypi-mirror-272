#
# GAMS - General Algebraic Modeling System Python API
#
# Copyright (c) 2017-2024 GAMS Development Corp. <support@gams.com>
# Copyright (c) 2017-2024 GAMS Software GmbH <support@gams.com>
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
#

import os
import warnings
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
)
import pandas as pd
from gams import transfer as gt
from gams.connect.agents.connectagent import ConnectAgent


class ExcelWriter(ConnectAgent):
    def __init__(self, system_directory, cdb, inst):
        super().__init__(system_directory, cdb, inst)
        try:
            opxl_version = list(map(int, openpyxl.__version__.split(".")))
            if opxl_version[0] < 3 or (opxl_version[0] == 3 and opxl_version[1] < 1):  # check openpyxl<3.1.0
              warnings.warn(f"The used openpyxl version is {openpyxl.__version__}. openpyxl<3.1.0 can lead to wrong data being exported. Upgrading to version 3.1.0 or later is recommended.", category=Warning)
        except ValueError:  # silence errors when converting to int
            pass
        self._parse_options()

        if self._trace > 3:
            pd.set_option("display.max_rows", None, "display.max_columns", None)

        if os.path.splitext(self._file)[1] != ".xlsx":
            self.connect_error("The ExcelWriter does support .xlsx files only.")

    def _parse_options(self):
        self._file = self._inst["file"]
        self._merged_cells = self._inst.get("mergedCells", False)
        self._cdim = self._inst.get("columnDimension", "auto")
        self._value_subs = self._inst.get("valueSubstitutions", None)
        self._clear_sheet = self._inst.get("clearSheet", False)
        self._trace = self._inst.get("trace", self._cdb.options.get("trace", 0))
        self._write_all = self._inst.get("writeAll", "auto")
        self._symbols = self._inst.get("symbols", [])

        if self._trace > 1:
            self._cdb.print_log(
                "Input (root):"
                f"\n  file: >{self._inst['file']}<"
                f"\n  columnDimension: >{self._inst.get('columnDimension', '')}<"
                f"\n  valueSubstitutions: >{self._inst.get('valueSubstitutions', '')}<"
                f"\n  clearSheet: >{self._inst.get('clearSheet', '')}<"
                f"\n  trace: >{self._inst.get('trace', '')}<"
                f"\n  writeAll: >{self._inst.get('writeAll', '')}<"
                "\n"
            )

        self._file = os.path.abspath(self._file)
        if self._write_all == "auto":
            self._write_all = True if self._symbols == [] else False

        if self._trace > 1:
            self._cdb.print_log(
                "Processed Input (root):"
                f"\n  file: >{self._file}<"
                f"\n  columnDimension: >{self._cdim}<"
                f"\n  valueSubstitutions: >{self._value_subs}<"
                f"\n  clearSheet: >{self._clear_sheet}<"
                f"\n  trace: >{self._trace}<"
                f"\n  writeAll: >{self._write_all}<"
                "\n"
            )

    def open(self):
        if os.path.exists(self._file):
            self._wb = openpyxl.load_workbook(
                self._file, read_only=False, data_only=False
            )
        else:
            self._wb = openpyxl.Workbook()
            # remove default sheet
            self._wb.remove(self._wb.active)

    def _write(self, df, rdim, cdim, sheet, nw_row, nw_col, merged_cells):
        row = nw_row
        use_index = True if rdim > 0 else False
        use_header = True if cdim > 0 else False
        last_row_labels = [None] * rdim

        for idx, r in enumerate(
            dataframe_to_rows(df, index=use_index, header=use_header)
        ):
            r2 = list(r)
            if idx < cdim:
                if merged_cells:
                    for idx2, x in enumerate(r2[rdim:]):
                        if idx2 == 0:  # first iteration
                            merge_start = merge_end = nw_col + rdim
                        elif x is None:
                            merge_end += 1
                            if idx2 == len(r2) - 1 - rdim:  # last iteration
                                if merge_start != merge_end:
                                    sheet.merge_cells(
                                        start_row=row,
                                        start_column=merge_start,
                                        end_row=row,
                                        end_column=merge_end,
                                    )
                        else:
                            if merge_start != merge_end:
                                sheet.merge_cells(
                                    start_row=row,
                                    start_column=merge_start,
                                    end_row=row,
                                    end_column=merge_end,
                                )
                            merge_start = merge_end = merge_end + 1
                else:
                    for idx2, x in enumerate(r2[rdim:]):
                        if x is None:
                            r2[idx2 + rdim] = r2[idx2 + rdim - 1]
            # remove extra empty row
            elif idx == cdim and rdim > 0:
                continue
            else:
                if merged_cells:
                    if idx == cdim + 1:  # first iteration
                        merge_start = [nw_row + cdim] * rdim
                        merge_end = [nw_row + cdim] * rdim
                    else:
                        for x in range(rdim):
                            if r2[x] is None:
                                merge_end[x] += 1
                                if idx == df.shape[0] + cdim:  # last iteration
                                    sheet.merge_cells(
                                        start_row=merge_start[x],
                                        start_column=x + nw_col,
                                        end_row=merge_end[x],
                                        end_column=x + nw_col,
                                    )
                            else:
                                if merge_start[x] != merge_end[x]:
                                    sheet.merge_cells(
                                        start_row=merge_start[x],
                                        start_column=x + nw_col,
                                        end_row=merge_end[x],
                                        end_column=x + nw_col,
                                    )
                                merge_start[x] = merge_end[x] = merge_end[x] + 1
                else:
                    for x in range(rdim):
                        if r2[x] is None:
                            r2[x] = last_row_labels[x]
                    last_row_labels = r2[:rdim]

            for col in range(len(r2)):
                if r2[col] is not None:
                    sheet.cell(row, col + nw_col).value = r2[col]
            row += 1

    def _pivot_cdim_only(self, df, dim, value_text):
        cols = df.columns.values.tolist()
        df["_first"] = value_text
        df = df.pivot(index=["_first"], columns=cols[:-1], values=[value_text])
        df.columns = df.columns.droplevel(0)  # remove column index "values"
        df.rename_axis([None], axis=0, inplace=True)  # remove index names
        df.rename_axis([None] * dim, axis=1, inplace=True)  # remove column index names
        return df

    def _pivot_rdim_only(self, df, dim, value_text):
        cols = df.columns.values.tolist()
        df["_first"] = value_text
        df = df.pivot(
            index=cols[:dim], columns=["_first"], values=[value_text]
        ).sort_index()
        df.columns = df.columns.droplevel(0)  # remove column index "values"
        df.rename_axis([None] * dim, axis=0, inplace=True)  # remove index names
        df.rename_axis([None], axis=1, inplace=True)  # remove column index names
        return df

    def _pivot_rdim_cdim(self, df, rdim, cdim, value_text):
        dim = rdim + cdim
        cols = df.columns.values.tolist()
        df = df.sort_values(by=cols[rdim:dim]).reset_index(drop=True)
        df = df.pivot(
            index=cols[:rdim], columns=cols[rdim:-1], values=[value_text]
        ).sort_index()
        df.columns = df.columns.droplevel(0)  # remove column index "values"
        df.rename_axis([None] * rdim, axis=0, inplace=True)  # remove index names
        df.rename_axis([None] * cdim, axis=1, inplace=True)  # remove column index names
        return df

    def _reshape_dataframe(self, df, dim, rdim, cdim, value):
        # turn integer column names into string column names to avoid problems reshaping if rdim=0
        df.columns = df.columns.map(str)
        if dim == 0:
            pass
        elif rdim == 0:
            df = self._pivot_cdim_only(df, dim, value)
        elif cdim == 0:
            df = self._pivot_rdim_only(df, dim, value)
        else:
            df = self._pivot_rdim_cdim(df, rdim, cdim, value)
        return df

    def _parse_range(self, sym_range, clear_sheet):
        def coords_to_row_col(coordinates):
            c = coordinate_from_string(coordinates)
            return c[1], column_index_from_string(c[0])

        if "!" not in sym_range:  # named range
            if sym_range not in self._wb.defined_names:
                self.connect_error(f"Named range >{sym_range}< does not exist.")
            rng = list(self._wb.defined_names[sym_range].destinations)
            if len(rng) > 1:
                self.connect_error(
                    f"Named ranges is not contiguous: {sym_range} -> {rng}"
                )
            rng = rng[0]
            resolved_range = "!".join(rng)
            sym_range = resolved_range
        elif sym_range.endswith("!"):  # sheet specified but range missing
            sym_range += "A1"

        sheet, rng = sym_range.split("!")
        for idx, s in enumerate(self._wb.sheetnames):
            if sheet.lower() == s.lower():
                if clear_sheet:
                    del self._wb[s]
                    sheet = self._wb.create_sheet(s, idx)
                else:
                    sheet = self._wb[s]
                break
        else:
            sheet = self._wb.create_sheet(sheet)

        nw, se = rng.split(":") if ":" in rng else (rng, None)

        nw_row, nw_col = coords_to_row_col(nw)
        se_col = se_row = None
        if se is not None:
            se_row, se_col = coords_to_row_col(se)

        return sheet, nw_col, nw_row, se_col, se_row

    def _expected_excel_shape(self, df, rdim, cdim):
        nr_rows = df.shape[0] + cdim
        nr_cols = df.shape[1] + rdim
        return (nr_rows, nr_cols)

    def _validate_range(self, df, rdim, cdim, nw_col, nw_row, se_col, se_row):
        required_rows, required_columns = self._expected_excel_shape(df, rdim, cdim)
        if se_row is not None:
            actual_rows = se_row - nw_row + 1
            if required_rows > actual_rows:
                return False
        if se_col is not None:
            actual_columns = se_col - nw_col + 1
            if required_columns > actual_columns:
                return False
        return True

    def _value_substitutions(self, df, value_subs, sym_type):
        vs = value_subs.copy() if value_subs is not None else {}
        if sym_type == "par":

            def replace_na_eps(df, eps_val, na_val):
                def isEps(x):
                    return False if isinstance(x, str) else gt.SpecialValues.isEps(x)

                def isNA(x):
                    return False if isinstance(x, str) else gt.SpecialValues.isNA(x)

                arr = df.iloc[:, -1].values.astype(object)
                eps_mask = isEps(arr)
                na_mask = isNA(arr)
                arr[eps_mask] = eps_val
                arr[na_mask] = na_val

                df[df.columns[-1]] = arr
                return df, eps_mask | na_mask

            # pandas does not distingish between GT special values NA and UNDEF - so we have to replace NA manually first
            # pandas replace() does replace 0 and -0 for key 0 or -0
            if "EPS" in vs.keys():
                eps_val = vs["EPS"]
                del vs["EPS"]
            else:
                eps_val = "EPS"
            if "NA" in vs.keys():
                na_val = vs["NA"]
                del vs["NA"]
            else:
                na_val = "NA"
            df, mask = replace_na_eps(df, eps_val, na_val)

            if "UNDEF" in vs.keys():
                vs[gt.SpecialValues.UNDEF] = vs["UNDEF"]
                del vs["UNDEF"]
            else:
                vs[gt.SpecialValues.UNDEF] = "UNDEF"
            if "INF" in vs.keys():
                vs[gt.SpecialValues.POSINF] = vs["INF"]
                del vs["INF"]
            else:
                vs[gt.SpecialValues.POSINF] = "INF"
            if "-INF" in vs.keys():
                vs[gt.SpecialValues.NEGINF] = vs["-INF"]
                del vs["-INF"]
            else:
                vs[gt.SpecialValues.NEGINF] = "-INF"
            if len(vs) > 0:
                df.iloc[~mask, -1] = df.iloc[~mask, -1].replace(vs)
        else:
            if len(vs) > 0:
                df.iloc[:, -1] = df.iloc[:, -1].replace(vs)
        return df

    def execute(self):
        try:
            if self._trace > 0:
                self.describe_container(self._cdb._container, "Connect Container")
            if self._write_all is True:
                self._symbols = [
                    {"name": s[0]}
                    for s in self._cdb._container
                    if isinstance(s[1], (gt.Parameter, gt.Set))
                ]
            for sym in self._symbols:
                if self._trace > 1:
                    self._cdb.print_log(
                        "Input (symbol):"
                        f"\n  name: >{sym.get('name', '')}<"
                        f"\n  range: >{sym.get('range', '')}<"
                        f"\n  mergedCells: >{sym.get('mergedCells', '')}<"
                        f"\n  columnDimension: >{sym.get('columnDimension', '')}<"
                        f"\n  valueSubstitutions: >{sym.get('valueSubstitutions', '')}<"
                        f"\n  clearSheet: >{sym.get('clearSheet', '')}<"
                        "\n"
                    )

                sym_name = sym.get("name")
                sym_range = sym.get("range", sym_name + "!A1")
                merged_cells = sym.get("mergedCells", self._merged_cells)
                cdim = sym.get("columnDimension", self._cdim)
                value_subs = sym.get("valueSubstitutions", self._value_subs)
                clear_sheet = sym.get("clearSheet", self._clear_sheet)

                if self._trace > 1:
                    self._cdb.print_log(
                        "Processed Input (symbol):"
                        f"\n name: >{sym_name}<"
                        f"\n range: >{sym_range}<"
                        f"\n mergedCells: >{merged_cells}<"
                        f"\n columnDimension: >{cdim}<"
                        f"\n valueSubstitutions: >{value_subs}<"
                        f"\n clearSheet: >{clear_sheet}<"
                        "\n"
                    )

                if sym_name not in self._cdb._container:
                    self.connect_error(
                        f"Symbol >{sym_name}< not found in Connect database."
                    )

                gt_sym = self._cdb._container[sym_name]

                if self._trace > 2:
                    self._cdb.print_log(
                        f"Connect Container symbol >{sym_name}<:\n {gt_sym.records}\n"
                    )

                if not isinstance(gt_sym, (gt.Set, gt.Parameter)):
                    self.connect_error(
                        f"Symbol type >{type(gt_sym)}< of symbol >{sym_name}< is not supported. Supported symbol types are set and parameter."
                    )
                sym_type = "par" if isinstance(gt_sym, gt.Parameter) else "set"

                if gt_sym.records is None or gt_sym.records.empty:
                    self._cdb.print_log(f"No data for symbol >{sym_name}<. Skipping.")
                    continue

                dim = gt_sym.dimension
                df = gt_sym.records.copy(deep=True)

                if isinstance(gt_sym, gt.Set):
                    value = "element_text"
                elif isinstance(gt_sym, gt.Parameter):
                    value = "value"

                if cdim == "auto":
                    cdim = 1 if dim > 0 else 0
                if self._trace > 2:
                    self._cdb.print_log(f"columnDimension: >{cdim}<")

                if cdim > dim:
                    self.connect_error(
                        f"columnDimension >{cdim}< exceeds dimension of symbol >{sym_name}<."
                    )
                rdim = dim - cdim

                df = self._value_substitutions(df, value_subs, sym_type)
                if self._trace > 2:
                    self._cdb.print_log(
                        f"DataFrame after valueSubstitutions ({sym_name}):\n{df}\n"
                    )

                if (
                    value == "element_text" and rdim * cdim > 0
                ):  # replace empty element_text by Y when exporting a true table
                    df.loc[df[value] == "", value] = "Y"

                sheet, nw_col, nw_row, se_col, se_row = self._parse_range(
                    sym_range, clear_sheet
                )

                df = self._reshape_dataframe(df, dim, rdim, cdim, value)
                if self._trace > 2:
                    self._cdb.print_log(
                        f"DataFrame after reshaping ({sym_name}):\n{df}\n"
                    )

                if not self._validate_range(
                    df, rdim, cdim, nw_col, nw_row, se_col, se_row
                ):
                    self.connect_error(f"Data exceeds range for symbol >{sym_name}<.")

                self._write(
                    df,
                    rdim,
                    cdim,
                    sheet,
                    nw_row,
                    nw_col,
                    merged_cells,
                )
        finally:
            self._wb.close()

    def close(self):
        if len(self._wb.sheetnames) == 0:
            self._cdb.print_log(f"No sheets in Excel file >{self._file}<. Skipping.")
        else:
            self._wb.save(self._file)
