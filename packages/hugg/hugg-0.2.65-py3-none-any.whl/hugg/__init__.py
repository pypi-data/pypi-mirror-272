import os,sys,types,importlib.machinery,shutil
import threading as thread
from pathlib import Path
from abc import ABC, abstractmethod
if sys.version_info[0] < 3: 
    from StringIO import StringIO
else:
    from io import StringIO
# https://pypi.org/project/ruamel.std.zipfile/
from ephfile import ephfile
from typing import List
from fileinput import FileInput as finput

fput = lambda foil, inplace=False, backup=None:finput(files=foil, inplace=inplace, backup=backup)

"""
class custom(mem):
    #https://docs.python.org/3/library/zipfile.html
    def __init__(self,wraplambda=lambda foil:False):
        super().__init__(wraplambda)

    def url(self):
        return

    def login(self):
        return
    
    def logout(self):
        return

    def files(self):
        return self.files
    
    def upload(self, file_path=None,path_in_repo=None):
        pass
    
    def download(self, file_path=None,download_to=None):
        pass
    
    def delete_file(self,path_in_repo=None):
        return False
"""
def isUTF(file):
    try:
        import codecs
        codecs.open(file, encoding="utf-8", errors="strict").readlines()
        return True
    except UnicodeDecodeError:
        return False


def split(a, n):
    """
    https://stackoverflow.com/questions/2130016/splitting-a-list-into-n-parts-of-approximately-equal-length

    >>> list(split(range(11), 3))
    [[0, 1, 2, 3], [4, 5, 6, 7], [8, 9, 10]]
    """
    k, m = divmod(len(a), n)
    return (a[i*k+min(i, m):(i+1)*k+min(i+1, m)] for i in range(n))

class mem(object):
    def __init__(self, wraplambda = lambda foil:False):
        self.dowraplambda = lambda foil:foil != None and wraplambda(foil) 

    @abstractmethod
    def url(self):
        pass

    @abstractmethod
    def login(self):
        pass
    
    @abstractmethod
    def logut(self):
        pass

    @abstractmethod
    def files(self):
        pass
    
    @abstractmethod
    def upload(self, path=None,path_in_repo=None):
        pass

    @abstractmethod
    def download(self, file_path=None,download_to=None):
        pass

    @abstractmethod
    def delete_file(self,path_in_repo=None):
        pass

    """
    def download(self, file_path=None,download_to=None):
        self.download(file_path, download_to)

        if self.dowraplambda(download_to) is True:
            download_to = self.unwrap(download_to)
        
        return download_to

    def upload(self, path=None,path_in_repo=None):
        from copy import deepcopy as dc
        og_path,og_path_in_repo = dc(path),dc(path_in_repo)
        if self.dowraplambda(path) is True:
            try:
                path = self.wrap(path)
                path_in_repo += ".nosj"
            except:
                path,path_in_repo = og_path,og_path_in_repo

        return self.upload(path, path_in_repo)
    """
    def __enter__(self):
        self.login()
        return self
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.logout()
        return self
    def __iadd__(self, path):
        self.upload(path)
        return self
    def __getitem__(self,foil):
        return self.download(foil)
    def __setitem__(self,key,value):
        self.upload(value,key)
    def __delitem__(self,item):
        return self.delete_file(item)
    def __str__(self):
        return self.files()
    def __contains__(self, item):
        return item in self.files()
    def __call__(self,item):
        return self.download(item) if item in self else None
    def __len__(self):
        return len(self.files())
    def concat(self,file,string):
        text = self.load_text(file)
        text += str("\n"+string)
        with open(file, "w+") as writer:
            writer.write(text)
        self.upload(file,file)
        try:os.remove(file)
        except:pass
    def outline(self):
        import pathlib
        output = {}
        for file in self.files():
            ext = pathlib.Path(file).suffix
            if ext not in output:
                output[ext] = 0
            output[ext] += 1
        return output
    def wrap(self,foil):
        if foil == None:
            return foil

        import json,mystring
        with open(foil,"r") as reader:
            content = reader.readlines()

        info = {
            'content':mystring.string(content).tobase64()
        }
        os.remove(foil)
        foil = foil+".nosj"

        with open(foil, "w+") as writer:
            writer.write(json.dumps(info))

        return foil

    def download_all(self, file_lambda, download_str_lambda=None):
        for foil in self.files():
            if file_lambda is not None and file_lambda(foil):
                self.download(
                    foil,
                    foil if download_str_lambda is None else download_str_lambda(foil)
                )

    def unwrap(self,foil):
        #Checking if there is a custom wrapping around the file, and unwrapping
        if foil != None and foil.endswith(".nosj"):
            import json,mystring
            with open(foil, 'r') as reader:
                content = json.load(reader)

            os.remove(foil)
            foil = foil.replace('.nosj','')

            with open(foil, 'w+') as writer:
                writer.write(
                    mystring.string.frombase64(content['contents'])
                )

        return foil


    def find_all(self,lambda_search,download:bool=False):
        return [self.download(x, os.path.basename(x)) if download else x for x in self.files() if lambda_search(x)]

    def logFiles(self, csvLogFileName:str):
        prep = lambda x:x.replace(',',';')
        with open(csvLogFileName, "w+") as writer:
            writer.write("FileNum,File,URL\n")
            for foilNum, foil in enumerate(self.files):
                writer.write("{}\n".format(','.join([
                    prep(foilNum), prep(foil),prep(self.url) 
                ])))
        return csvLogFileName

    def by(self,ext,download:bool=False):
        return self.find_all(lambda_search=lambda x:str(x).endswith(ext), download=download)

    def find(self,lambda_search,download:bool=False):
        current = self.find_all(lambda_search,download)
        if len(current) > 1:
            print("There are too many files found")
        elif len(current) == 1:
            return current[0]
        return None

    def impor(self,file,delete=False):
        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return None
        
        file_path = self[file]
        import_name = str(file.split('/')[-1]).replace('.py','')
        #https://stackoverflow.com/questions/19009932/import-arbitrary-python-source-file-python-3-3#answer-19011259
        loader = importlib.machinery.SourceFileLoader(import_name, os.path.abspath(file_path))
        mod = types.ModuleType(loader.name)
        loader.exec_module(mod)
        if delete:
            os.remove(file_path)

        return mod

    def load_by_ext(self, ext:str):
        formatting = {
            ".json":self.load_json,
            ".csv":self.load_csv,
            ".pkl":self.load_pkl,
            ".sqlite":self.load_sqlite,
            ".xlsx":self.load_xcyl,
            ".zip":self.load_zip,
            ".json":self.load_json,
        }
        if ext is None:
            return None

        if ext not in formatting:
            return self.try_load

        return formatting[ext]

    def load_text(self,file):
        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return None
        
        cur_path = os.path.abspath(self[file])
        with open(cur_path, 'r') as reader:
            contents = reader.readlines()
        os.remove(cur_path)

        return ''.join(contents)

    def load_binary(self,file):
        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return None
        
        cur_path = os.path.abspath(self[file])
        with open(cur_path, 'rb') as reader:
            contents = reader.read()
        os.remove(cur_path)

        return contents
    
    def try_load(self,file):
        output = {"contents":None,"isbinary":None}

        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return output

        try:
            contents = self.load_text(file)
            output["contents"] = contents
            output["isbinary"] = False
        except:pass

        try:
            contents = self.load_binary(file)
            output["contents"] = contents
            output["isbinary"] = True
        except:pass

        return output

    def load_json(self,file):
        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return None
        
        import json
        
        cur_path = os.path.abspath(self[file])
        with open(cur_path, 'r') as reader:
            contents = json.load(reader)
        os.remove(cur_path)

        return contents

    def load_pkl(self,file):
        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return None

        import pandas as pd

        cur_path = os.path.abspath(self[file])
        contents = pd.read_pickle(cur_path)
        os.remove(cur_path)

        return contents

    def load_csv(self,file):
        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return None

        import pandas as pd

        cur_path = os.path.abspath(self[file])
        contents = pd.read_csv(cur_path)
        os.remove(cur_path)

        return contents

    def load_split_csv(self,prefix):
        if not prefix.endswith("*"):
            prefix = prefix + "*"

        import pandas as pd
        current_data = []
        
        for foil in self.files():
            import fnmatch
            if fnmatch.fnmatch(foil, prefix):
                current_file = os.path.abspath(self[foil])

                with open(current_file,'r') as reader:
                    current_data += reader.readlines()
                
                os.remove(current_file)

        return pd.read_csv(StringIO('\n'.join(current_data)))

    def load_sqlite(self,file):
        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return None
        
        import xcyl
        
        cur_path = os.path.abspath(self[file])
        contents = xcyl.sqlobj(cur_path)

        return contents

    def load_xcyl(self,file):
        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return None
        
        import xcyl
        
        cur_path = os.path.abspath(self[file])
        contents = xcyl.xcylobj(cur_path)

        return contents

    def load_zip(self,file):
        if file not in self.files():
            print("FILE IS NOT AVAILABLE")
            return None

        return zyp(os.path.abspath(self[file]))

    def empty(self,save_files=['README.md','.gitattributes','.gitignore']):
        print('[',end='',flush=True)
        filez = [x for x in self.files() if x not in save_files]
        _range = range(0, len(filez))

        class MyThread(thread.Thread):
            #https://www.tutorialspoint.com/python3/python_multithreading.htm
            def __init__(self,my_range,delete_file_lambda):
                thread.Thread.__init__(self)
                self.my_range = my_range
                self.delete_file_lambda = delete_file_lambda
            def run(self):
                for num in self.my_range:
                    self.delete_file_lambda(num)
                    print('.',end='',flush=True)
                    #delete_msg(self.name,self.bot,self.chat_id,num)

        thread_dyct = []
        ranges = split(_range,self.threads)
        for itr,thread_range in enumerate(ranges):
            thread_dyct += [
                MyThread(thread_range, lambda x:self.delete_file(x))
            ]
        [tred.start() for tred in thread_dyct]
        [tred.join() for tred in thread_dyct]

        """
        for foil in self.files():
            if foil not in save_files:
                self.delete_file(foil)
        """
        print(']')
    
    def file_ext_metrics(self, full_path_lambda=None):
        extensions = {}
        for foil in self.files():
            if "." in foil and (full_path_lambda is None or full_path_lambda(foil)):
                ext = foil.split(".")[-1]
                if ext not in extensions:
                    extensions[ext] = 0
                extensions[ext] += 1
        return extensions
    
    def file_size_metrics(self):
        file_sizes = {}
        for foil in self.files():
            if "." in foil and not foil.startswith("."):
                file_sizes[foil] = os.path.getsize(foil)
        return file_sizes
    
    def file_listing_write_out(self, csv_file_name:str):
        from pathlib import Path
        with open(csv_file_name, "w+") as writer:
            writer.write("FileNum, FileSize(pathlib stat st_size), FileExt, FilePath\n")
            for foil_itr, foil in enumerate(self.files()):
                foil = Path(foil)
                writer.write(", ".join([
                    foil_itr,
                    foil.stat().st_size,
                    foil.suffix,
                    foil.absolute()
                ]))
    
    @staticmethod
    def from_splych(core_file_name):
        try:
            import splych
            full_file = splych.file_stitch(core_file_name)
            split = os.path.splitext(full_file)

            if len(split) > 1 and split[-1] in available_types().keys():
                return available_types()[split[-1]]()
            else:
                return full_file
        except:
            return None

    @staticmethod
    def to_splych(file_name, chunks_size=None, delete_og=False):
        try:
            import splych
            return file_split(
                file=file_name,
                parts=None,
                chunk_size=chunks_size,
                delete_original=delete_og,
                generate_hashfile=False,
                generate_config=False
            )
        except:
            return []

    def filesystem(self):
        #https://www.pyfilesystem.org/
        #https://stackoverflow.com/questions/51508179/how-to-construct-an-in-memory-virtual-file-system-and-then-write-this-structure
        import fs
        import fs.copy
        from fs.tempfs import TempFS
        mem_fs = fs.open_fs("temp://")
        for file in self.files():
            try: mem_fs.makedirs(os.path.dirname(file))
            except:pass

            contents = None
            try: contents = self.try_load(file)
            except:pass

            if contents is not None and contents['contents'] is not None:
                with mem_fs.open(file, "wb+" if contents['isbinary'] else "w+") as writer:
                    writer.write(contents['contents'])
        return mem_fs

class localdrive(mem):
    #https://python-gitlab.readthedocs.io/en/stable/index.html#installation
    #https://python-gitlab.readthedocs.io/en/stable/api-usage.html
    def __init__(self,path:str=os.curdir,wraplambda=lambda foil:False):
        super().__init__(wraplambda)
        self.path = path

    def files(self):
        return [os.path.join(dp, f) for dp, dn, filenames in os.walk(self.path) for f in filenames if os.path.isfile(f)]

    def login(self):
        return
    
    def logout(self):
        return
    
    def download(self, file_path=None,download_to=None):
        download_to = download_to or os.path.basename(file_path)
        shutil.copy(file_path, download_to)
        return download_to
    
    def upload(self, file_path=None,path_in_repo=None):
        shutil.copy(file_path, path_in_repo)
        return True
    
    def delete_file(self,path_in_repo=None):
        if path_in_repo in self.files():
            os.remove(path_in_repo)
        return True

class subRemote(mem):
    #https://docs.python.org/3/library/zipfile.html
    def __init__(self,identifyFile:str,wraplambda=lambda foil:False):
        super().__init__(wraplambda)
        self.identifyFile = identifyFile
        __name__ = ''

        self.files = []
        self.download = None

        if os.path.exists(self.identifyFile):
            self.mod = __import__(self.identifyFile)
            self.download = self.mod.download
            self.files = self.mod.info()

    def url(self):
        self.location = zyp_file

    def login(self):
        return
    
    def logout(self):
        return

    def files(self):
        return self.files
    
    def upload(self, file_path=None,path_in_repo=None):
        return False
    
    def download(self, file_path=None,download_to=None):
        if self.download == None:
            return False
        self.download(file_path, download_to)
    
    def delete_file(self,path_in_repo=None):
        return False

try:
    from huggingface_hub import HfApi
    class face(mem):
        def __init__(self,repo,use_auth=True,repo_type="dataset",clear_cache=False, clear_token=False,wraplambda=lambda foil:False):
            super().__init__(wraplambda)
            """
            https://rebrand.ly/hugface

            https://huggingface.co/docs/huggingface_hub/quick-start
            https://huggingface.co/docs/huggingface_hub/how-to-upstream
            https://huggingface.co/docs/huggingface_hub/how-to-downstream
            """
            self.api = HfApi()
            self.repo = repo
            self.repo_type = repo_type
            self.auth = use_auth
            self.downloaded_files = []
            self.opened = False
            self.clear_cache = clear_cache
            self.clear_token = clear_token
            self._pr_ = {}
            self.backup_auth = use_auth
            self.threads = 4

        def get_pull_requests(self, status='open'):
            #https://huggingface.co/docs/huggingface_hub/how-to-discussions-and-pull-requests#retrieve-discussions-and-pull-requests-from-the-hub
            #https://github.com/huggingface/huggingface_hub/blob/v0.9.0/src/huggingface_hub/hf_api.py#L2475
            if not self.opened:
                self.login()

            output = []

            try:
                output = [
                    x for x in self.api.get_repo_discussions(repo_id=self.repo,repo_type=self.repo_type,token=self.auth)
                    if x.is_pull_request and x.status==status
                ]
            except Exception as e:
                print(e)

            return output
        
        def get_pull_request_info(self, pull_request_num):
            #https://github.com/huggingface/huggingface_hub/blob/v0.9.0/src/huggingface_hub/hf_api.py#L2554
            output = None
            if not self.opened:
                self.login()

            try:
                output = self.api.get_discussion_details(repo_id=self.repo,repo_type=self.repo_type,token=self.auth, discussion_num=pull_request_num)
            except Exception as e:
                print(e)

            return output
        
        def url(self):
            return "https://huggingface.co/datasets/" + str(self.repo)

        @property
        def pr(self):
            if self._pr_ == {}:
                for pr in self.get_pull_requests():
                    print(pr)
                    self._pr_[pr.num] = pr
            return self._pr_

        def merge_pull_request(self, discussion_id=-1, comment="Auto Merge of the Pull Request"):
            #https://huggingface.co/docs/huggingface_hub/v0.9.0/en/package_reference/hf_api#huggingface_hub.HfApi.merge_pull_request
            #https://github.com/huggingface/huggingface_hub/blob/v0.9.0/src/huggingface_hub/hf_api.py#L3033
            if not self.opened:
                self.login()

            try:
                self.api.merge_pull_request(
                    repo_id=self.repo,
                    discussion_num=discussion_id,
                    comment=comment,
                    repo_type=self.repo_type,
                    token=self.backup_auth or self.auth
                )
                if discussion_id in self._pr_:
                    del self._pr_[discussion_id]
                return True
            except Exception as e:
                print(e)
                return False

        def merge_pull_requests(self, comment="Auto Merge of the Pull Request"):
            #https://huggingface.co/docs/huggingface_hub/v0.9.0/en/package_reference/hf_api#huggingface_hub.HfApi.merge_pull_request
            #https://github.com/huggingface/huggingface_hub/blob/v0.9.0/src/huggingface_hub/hf_api.py#L3033
            """ MULTITHREAD PREP
            threads = 4
            try:
                delete_to = update.message.message_id+1
                delete_from = last_message(delete_to)

                _range = range(delete_from, delete_to)
                print(f"Deleting from {delete_from} to {delete_to}")

                class MyThread(thread.Thread):
                    #https://www.tutorialspoint.com/python3/python_multithreading.htm
                    def __init__(self,name,bot,chat_id,msg_range):
                        thread.Thread.__init__(self)
                        self.name = name
                        self.bot = bot
                        self.chat_id = chat_id
                        self.msg_range = msg_range
                    def run(self):
                        for num in self.msg_range:
                            delete_msg(self.name,self.bot,self.chat_id,num)
                
                thread_dyct = []
                ranges = split(_range,threads)
                print(ranges)
                for itr,thread_range in enumerate(ranges):
                    print(f"Range:> {thread_range}")
                    thread_dyct += [
                        MyThread(itr, context.bot, update.message.chat_id, thread_range)
                    ]
                [tred.start() for tred in thread_dyct]
                [tred.join() for tred in thread_dyct]
            except Exception as e:
                update.message.reply_text(f"failure: {e}")
            """

            if not self.opened:
                self.login()

            try:
                for pr in self.get_pull_requests():
                    self.merge_pull_request(
                        discussion_id=pr.num,
                        comment=comment
                    )
            except Exception as e:
                print(e)

            return None

        def clearcache(self, force=False, all_sets=False):
            """ MULTITHREAD PREP
            threads = 4
            try:
                delete_to = update.message.message_id+1
                delete_from = last_message(delete_to)

                _range = range(delete_from, delete_to)
                print(f"Deleting from {delete_from} to {delete_to}")

                class MyThread(thread.Thread):
                    #https://www.tutorialspoint.com/python3/python_multithreading.htm
                    def __init__(self,name,bot,chat_id,msg_range):
                        thread.Thread.__init__(self)
                        self.name = name
                        self.bot = bot
                        self.chat_id = chat_id
                        self.msg_range = msg_range
                    def run(self):
                        for num in self.msg_range:
                            delete_msg(self.name,self.bot,self.chat_id,num)
                
                thread_dyct = []
                ranges = split(_range,threads)
                print(ranges)
                for itr,thread_range in enumerate(ranges):
                    print(f"Range:> {thread_range}")
                    thread_dyct += [
                        MyThread(itr, context.bot, update.message.chat_id, thread_range)
                    ]
                [tred.start() for tred in thread_dyct]
                [tred.join() for tred in thread_dyct]
            except Exception as e:
                update.message.reply_text(f"failure: {e}")
            """
            if self.clear_cache or force:
                cache_loc = "/home/"+str(os.getlogin()) + "/.cache/huggingface/hub/"
                user_name, repo_name = self.repo.split('/')

                paths_to_remove = [
                    str(cache_loc) + "datasets--{0}--{1}/".format(user_name,repo_name)
                ]

                if all_sets:
                    paths_to_remove += [
                        str(cache_loc + x) for x in os.listdir(cache_loc)
                    ]

                for y in paths_to_remove:
                    try:
                        os.system("yes|rm -r " + str(y))
                    except:
                        pass

        def login(self):
            if isinstance(self.auth,str):
                import os

                hugging_face = os.path.join(Path.home(),".huggingface")
                token_path = os.path.join(hugging_face, "token")

                if os.path.exists(token_path) and self.clear_token:
                    os.system("rm {0}".format(token_path))

                if not os.path.exists(token_path):
                    for cmd in [
                        f"mkdir -p {hugging_face}",
                        f"rm {token_path}",
                        f"touch {token_path}"
                    ]:
                        try:
                            os.system(cmd)
                        except:
                            pass

                    with open(token_path,"a") as writer:
                        writer.write(self.auth)
                self.auth = True
            self.clearcache()
            self.opened = True
            return

        def logout(self):
            for foil in self.downloaded_files:
                try:
                    os.remove(foil)
                except:
                    try:
                        os.system("yes|rm " + str(foil))
                    except Exception as e:
                        print("Failed to remove the cached file " +str(foil))
                        print(e)
                        pass
            self.clearcache()
            return

        def download(self, file_path=None,download_to=None):
            download_to = download_to or os.path.basename(file_path)
            if not self.opened:
                self.login()
            #https://huggingface.co/docs/huggingface_hub/v0.9.0/en/package_reference/file_download#huggingface_hub.hf_hub_download
            if file_path and isinstance(file_path,str):
                from huggingface_hub import hf_hub_download
                current_file = hf_hub_download(
                    repo_id=self.repo,
                    filename=file_path,
                    repo_type=self.repo_type,
                    use_auth_token=self.auth
                )
                if download_to:
                    try:
                        shutil.copy(current_file, os.path.basename(current_file))
                        current_file = os.path.basename(current_file)
                    except:
                        pass
                return current_file
            return None

        def close_nonchanging_pr(self, pull_request=None):
            """ MULTITHREAD PREP
            threads = 4
            try:
                delete_to = update.message.message_id+1
                delete_from = last_message(delete_to)

                _range = range(delete_from, delete_to)
                print(f"Deleting from {delete_from} to {delete_to}")

                class MyThread(thread.Thread):
                    #https://www.tutorialspoint.com/python3/python_multithreading.htm
                    def __init__(self,name,bot,chat_id,msg_range):
                        thread.Thread.__init__(self)
                        self.name = name
                        self.bot = bot
                        self.chat_id = chat_id
                        self.msg_range = msg_range
                    def run(self):
                        for num in self.msg_range:
                            delete_msg(self.name,self.bot,self.chat_id,num)
                
                thread_dyct = []
                ranges = split(_range,threads)
                print(ranges)
                for itr,thread_range in enumerate(ranges):
                    print(f"Range:> {thread_range}")
                    thread_dyct += [
                        MyThread(itr, context.bot, update.message.chat_id, thread_range)
                    ]
                [tred.start() for tred in thread_dyct]
                [tred.join() for tred in thread_dyct]
            except Exception as e:
                update.message.reply_text(f"failure: {e}")
            """
            #https://huggingface.co/docs/huggingface_hub/main/en/package_reference/hf_api#huggingface_hub.HfApi.change_discussion_status
            for pr in self.get_pull_requests():
                if pull_request is None or pull_request == pr.num:
                    pr_details = self.get_pull_request_info(pr.num)
                    if pr_details.diff.strip() == '':
                        self.api.change_discussion_status(
                            repo_id=self.repo,
                            repo_type=self.repo_type,
                            token=self.auth,
                            discussion_num=pr.num,
                            new_status='closed'
                        )

        def upload(self, path=None,path_in_repo=None, auto_accept_all_pull_requests=True,use_pull_request = True):
            if not self.opened:
                self.login()
            #Because HuggingFace_hub will get pissed if we don't use it
            if path:
                if isinstance(path,str) and os.path.isfile(path):
                    #https://huggingface.co/docs/huggingface_hub/v0.9.0/en/package_reference/hf_api#huggingface_hub.HfApi.upload_file
                    self.api.upload_file(
                        path_or_fileobj=path,
                        path_in_repo=path_in_repo or path,
                        repo_id=self.repo,
                        repo_type=self.repo_type,
                        create_pr=use_pull_request #https://huggingface.co/docs/huggingface_hub/v0.10.0.rc0/en/how-to-discussions-and-pull-requests
                    )
                    if auto_accept_all_pull_requests:
                        self.merge_pull_requests()
                elif isinstance(path,str) and os.path.isdir(path):
                    #https://huggingface.co/docs/huggingface_hub/v0.9.0/en/package_reference/hf_api#huggingface_hub.HfApi.upload_folder
                    self.api.upload_file(
                        folder_path=path,
                        path_in_repo=path_in_repo or path,
                        repo_id=self.repo,
                        repo_type=self.repo_type,
                        create_pr=use_pull_request
                    )
                    if auto_accept_all_pull_requests:
                        self.merge_pull_requests()
                else:
                    print("Entered path " + str(path) + " is not supported or doesn't exist exists(" +  str(os.path.exists(path)) + ").")
                return True
            return False

        def files(self):
            if not self.opened:
                self.login()
            # https://huggingface.co/docs/huggingface_hub/v0.9.0/en/package_reference/hf_api#huggingface_hub.HfApi.list_repo_files
            return self.api.list_repo_files(
                repo_id=self.repo,
                repo_type=self.repo_type
            )

        def ol_impor(self,file):
            if file not in self.files():
                print("FILE IS NOT AVAILABLE")
                return None
            
            import_name = str(file.split('/')[-1]).replace('.py','')
            #https://stackoverflow.com/questions/19009932/import-arbitrary-python-source-file-python-3-3#answer-19011259
            loader = importlib.machinery.SourceFileLoader(import_name, os.path.abspath(self[file]))
            mod = types.ModuleType(loader.name)
            loader.exec_module(mod)

            return mod
            
        def delete_file(self,path_in_repo=None,use_flag=False):
            if not self.opened:
                self.login()
            # https://huggingface.co/docs/huggingface_hub/v0.9.0/en/package_reference/hf_api#huggingface_hub.HfApi.delete_file
            if path_in_repo:
                self.api.delete_file(
                    path_in_repo=path_in_repo,
                    repo_id=self.repo,
                    repo_type=self.repo_type,
                    create_pr=use_flag
                )
                if use_flag:
                    self.merge_pull_requests()
            return False
            
        def to_ghub(self, location, access_token):
            if not self.opened:
                self.login()

            ghub_repo = ghub(location, access_token, create=True)

            for foil in self.files():
                ghub_repo[foil] = self[foil]

            return ghub_repo



    class fixface(face):
        @staticmethod
        def run(cmd):
            print(cmd);os.system(cmd)

        def __init__(self,repo,use_auth=True,repo_type="dataset",clear_cache=False, clear_token=False, sparse=False,wraplambda=lambda foil:False):
            super().__init__(repo,use_auth=True,repo_type="dataset",clear_cache=False, clear_token=False,wraplambda=wraplambda)
            #https://github.blog/2020-01-17-bring-your-monorepo-down-to-size-with-sparse-checkout/
            fixface.run("git clone {1} https://huggingface.co/datasets/{0}".format(repo, "--no-checkout" if sparse else ""))
            if sparse:
                fixface.run("cd {0} && git sparse-checkout init --cone".format(repo.split("/")[-1]))

        def __enter__(self):
            return self
        
        def exit(self):
            fixface.run("yes|rm -r {0}/".format(self.repo.split("/")[-1]))

        def __exit__(self,exc_type, exc_val, exc_tb):
            self.exit()
            return self

        def fix_pr(self, num):
            num = str(num)
            def run(cmd):
                print(cmd);os.system(cmd)

            run("git fetch origin refs/pr/{0}:pr/{0}".format(num))

            class pr(object):
                def __init__(self,num,face=None):
                    self.num = num
                    self.face = face
                    run("cd {0}".format(repo.split("/")[-1]))
                def fixattr(self):
                    run("git checkout main -- .gitattributes && git add .gitattributes")
                def __enter__(self):
                    run("git checkout pr/{0}".format(self.num))
                    return self
                def __exit__(self,exc_type, exc_val, exc_tb):
                    run("git commit -m \"Fixed the gitattributes\"")
                    run("git push origin pr/{0}:refs/pr/{0}".format(self.num))
                    run("git checkout main")
                    face.merge_pull_request(self.num)
                    return self
            return pr(num,self)
except: pass

try:
    from github import Github,Requester
    import gett as wget
    from mystring import gh_url as githuburl
    #https://pygithub.readthedocs.io/en/latest/
    #https://pygithub.readthedocs.io/en/latest/examples/Branch.html#get-a-branch
    class ghub(mem):
        @staticmethod
        def create_repo(auth_key, repo_name, private=True):
            req = Requester.Requester(auth_key,None,None,"https://api.github.com",15,"PyGithub/Python",30,True,None,None)
            try:
                headers, data = req.requestJsonAndCheck(
                    "POST", "https://api.github.com/user/repos", parameters={},headers={
                        "Accept":"application/vnd.github+json",
                        "Authorization":"Bearer {0}".format(auth_key),
                        "X-GitHub-Api-Version":"2022-11-28",
                    }, input = {
                        "name":repo_name,
                        "private":private,
                        "auto_init":True,
                    }
                )
                output = True
            except:
                output = False

            return output

        def apiURL(self, filePath:str=None):
            url = 'https://api.github.com/repos/{repo}'.format(repo=self.baserepo)
            if filePath is not None:
                url += "/contents/{path}".format(path=filePath)
            return url

        def __init__(self,repo,access_token,branch_commitsha='master',usewget=False, branch='master',create=False,wraplambda=lambda foil:False, timeout:int=60 * 10, retries:int=15):
            super().__init__(wraplambda)

            self.token = access_token
            self.baserepo = repo
            self.repo = repo
            self.github_access = Github(access_token, timeout=timeout)
            self.usewget = usewget
            self.retries = retries
            if create:
                try:
                    repo = self.github_access.get_repo(repo)
                    has_repo = True
                except:
                    has_repo = False
                
                if has_repo:
                    raise Exception("There already is an repo with this name")
                
                if not ghub.create_repo(access_token, repo):
                    raise Exception("Error creating repo")

            if False:
                #create
                #https://docs.github.com/en/rest/repos/repos?apiVersion=2022-11-28#create-a-repository-for-the-authenticated-user

                self.github_access = Github(access_token)

                #search :> https://github.com/PyGithub/PyGithub/blob/master/github/MainClass.py#L410
                #https://docs.github.com/en/rest/repos/repos?apiVersion=2022-11-28#create-an-organization-repository
                if create:
                    print('a')
                else:
                    self.repo = self.github_access.get_repo(repo)
        
            self.repo = self.github_access.get_repo(repo)

            self.branch_commitsha = None
            #https://stackoverflow.com/questions/59148874/get-all-the-file-contents-of-a-repo-at-specific-commit-using-pygithub
            if branch_commitsha is not None:
                self.branch_commitsha = branch_commitsha
            elif branch is not None:
                self.branch_commitsha = branch
            else:
                self.branch_commitsha = "master"
            
            self.ghub_url = githuburl(self.repo, token=self.token)

        def files(self):
            files = []
            contents = self.repo.get_contents("", ref=self.branch_commitsha)
            while contents:
                file_content = contents.pop(0)
                if file_content.type == "dir":
                    contents.extend(self.repo.get_contents(file_content.path, ref=self.branch_commitsha))
                else:
                    files += [file_content.path]
            return files

        def url(self):
            return "https://github.com/" + str(self.repo.full_name)

        def login(self):
            return
        
        def logout(self):
            return
        
        def has_repo(self, repo):
            try:
                repo = self.github_access.get_repo(repo)
                output = True
            except:
                output = False
            return output
        
        def download(self, file_path=None,download_to=None, encoding="utf-8"):
            #download_to = download_to or os.path.basename(file_path)
            if download_to is None:
                download_to = os.path.join(os.curdir,file_path.split("/")[-1])
            if file_path and isinstance(file_path,str):
                if not os.path.exists(download_to):
                    from pathlib import Path
                    filePath = Path(download_to)
                    try:
                        os.makedirs(filePath.parent.absolute())
                    except: pass
                    #try:
                    #    filePath.touch()
                    #except: pass

                if self.usewget:
                    download_from = "{0}/raw/{1}/{2}".format(self.url(), self.branch_commitsha, file_path)
                    attempt_ktr = 0
                    while attempt_ktr != self.retries:
                        try:
                            wget.download(url=download_from, out=download_to)
                            attempt_ktr = self.retries
                        except Exception as e:
                            print("Error := {0}".format(e))
                            print("Attempt[{0}] >= {1}".format(attempt_ktr, download_from))
                            print(str(self.repo.full_name))
                            print(file_path)
                            attempt_ktr += 1
                else:
                    current_contents = self.repo.get_contents(file_path, ref=self.branch_commitsha)

                    if encoding is not None and encoding.strip() != '' and not file_path.endswith('.zip'):
                        current_contents = current_contents.decoded_content.decode(encoding)
                        encode_writing = "w+"
                    else:
                        import base64
                        #base64.b64decode(bytearray(self.content, "utf-8"))
                        #print(current_contents.content)
                        #print(bytearray(current_contents.content, "utf-8"))
                        #print(base64.b64decode(bytearray(current_contents.content, "utf-8")))
                        current_contents = base64.b64decode(bytearray(current_contents.content, "utf-8"))
                        encode_writing = "wb+"

                    with open(download_to,encode_writing) as writer:
                        writer.write(current_contents)
            return download_to
        
        def upload(self, file_path=None,path_in_repo=None,debug:bool=False):
            if isUTF(file_path):
                from pathlib import Path
                new_contents = Path(file_path).read_text()

                if path_in_repo in self: #Update
                    contents = self.repo.get_contents(path_in_repo, ref=self.branch_commitsha) #https://github.com/PyGithub/PyGithub/blob/001970d4a828017f704f6744a5775b4207a6523c/github/Repository.py#L1803
                    self.repo.update_file(contents.path, "Updating the file {}".format(path_in_repo), new_contents, contents.sha, branch=self.branch_commitsha) #https://github.com/PyGithub/PyGithub/blob/001970d4a828017f704f6744a5775b4207a6523c/github/Repository.py#L2134
                else: #Create #https://github.com/PyGithub/PyGithub/blob/001970d4a828017f704f6744a5775b4207a6523c/github/Repository.py#L2074
                    self.repo.create_file(path_in_repo, "Creating the file {}".format(path_in_repo), new_contents, branch=self.branch_commitsha)
            else:
                import base64, requests
                with open(file_path, "rb") as f:
                    r = requests.put(self.apiURL(path_in_repo), headers={
                        "Authorization": "Bearer {0}".format(self.token),
                        "Content-type": "application/vnd.github+json"
                    }, json={
                        "message": "Creating/Updating the file {}".format(path_in_repo),
                        "content": base64.b64encode(f.read()).decode("utf-8")
                    })
                    if debug:
                        print(r.text)

        def delete_file(self,path_in_repo=None):
            if path_in_repo in self.files():
                contents = self.repo.get_contents(path_in_repo, ref=self.branch_commitsha) #https://github.com/PyGithub/PyGithub/blob/001970d4a828017f704f6744a5775b4207a6523c/github/Repository.py#L1803
                self.repo.delete_file(path_in_repo, "Deleting the file {}".format(path_in_repo), contents.sha,branch=self.branch_commitsha) #https://github.com/PyGithub/PyGithub/blob/001970d4a828017f704f6744a5775b4207a6523c/github/Repository.py#L2198
except:pass

try:
    import threading
    GRepo_Saving_Progress_Lock = threading.Lock()

    from github import Github
    import gett as wget
    import os, asyncio
    import os, requests, datetime, time, queue, asyncio
    from copy import deepcopy as dc
    from threading import Lock
    from typing import Dict, List, Callable, Generic, TypeVar
    from abc import ABC, abstractmethod
    from fileinput import FileInput as finput
    import mystring,splittr
    import pause
    import pygit2 as git2
    from contextlib import suppress

    T = TypeVar('T')
    class GRepo_Seed_Metric(ABC, Generic[T]):
        @abstractmethod
        def name(self) -> str:
            pass

        @abstractmethod
        def metric(self, filename: str, source_code: str) -> T:
            pass

        @abstractmethod
        def diff(self, latest: T, previous: T):
            pass

        def __call__(self):
            setattr(self.metric, 'diff', self.diff)
            return self.metric

    class q_ghub(object):
        def __init__(self,token,query_string:str,usewget=False, metrics:List[GRepo_Seed_Metric]=[], num_processes:int = None, start_num:int=-1,end_num:int=-1,delete_paths:bool=False, num_threads:int=10):
            self.token = token
            self.query_string = query_string
            
            self.metrics = metrics
            self.token = token
            if "GH_TOKEN" not in os.environ:
                self.login()

            self.usewget = usewget

            self.g = Github(self.token)
            self.processor = mystring.MyThreads(num_threads)
            self.processed_paths = queue.Queue()
            setattr(self.processed_paths, 'lock', threading.Lock())

            self.current_repo_itr = None
            self.total_repo_len = None
            self.num_processes = num_processes

            def appr(string: mystring.string):
                with open("mapping_file_{0}.csv".format(string.tobase64()), "a+") as writer:
                    writer.write(string)
            self.appr = appr
            self.api_watch = mystring.gh_api_status()
            self.delete_paths = delete_paths
            self.tracking_repos = None
            self.tracking_name = None

            self.start_num = start_num
            self.end_num = end_num

            #Seems Sus
            asyncio.run(self.handle_history())

        def login(self):
            os.environ['GH_TOKEN'] = self.token
            with suppress(Exception):
                with open("~/.bashrc", "a+") as writer:
                    writer.write("GH_TOKEN={0}".format(self.token))

        def save(self, current_project_url:str=None):
            with GRepo_Saving_Progress_Lock:
                if not os.path.exists(self.localfilename):
                    with open(self.localfilename, "w+") as writer:
                        writer.write("ProjectItr,ProjectURL,ProjectScanned\n")
                        for proj_itr, proj in enumerate(self.repos):
                            writer.write("{0},{1},false\n".format(proj_itr, proj))

                if current_project_url is not None:
                    found = False
                    with finput(self.localfilename, inplace=True) as reader:
                        for line in reader:
                            if not found and current_project_url in line:
                                line = line.replace("false", "true")
                            print(line, end='')
            return

        async def handle_history(self):
            while not self.complete:
                # Get up to 5 strings from the queue
                paths,num_waiting = [], 5
                while len(paths) < num_waiting:
                    try:
                        path = self.processed_paths.get()
                        paths.append(path)
                    except queue.Empty:
                        time.sleep(10)
                        num_waiting -= 1

                # Save Externally Here
                for path in paths:
                    string("yes|rm -r {0}".format(path)).exec()

                for path in paths:
                    self.save(path)

        def mine_repo(self, repo_dir, sqlite_db_file):
            import git4net as git2net
            if self.num_processes is None:
                git2net.mine_git_repo(repo_dir, sqlite_db_file)
            else:
                git2net.mine_git_repo(repo_dir, sqlite_db_file, no_of_processes=self.num_processes)

            git2net.mining_state_summary(repo_dir, sqlite_db_file)
            git2net.disambiguate_aliases_db(sqlite_db_file)
            git2net.compute_complexity(repo_dir, sqlite_db_file, no_of_processes=1, extra_eval_methods=self.metrics)

        def mine_repos(self):
            self.timing

            def process_prep(repo_itr:int, repo_clone_url:str, search_string:str, appr:Callable, fin_queue:queue.Queue):
                self.query_string = search_string
                def process():
                    def repair(path,create:bool=True, delete_paths:bool=False):
                        if delete_paths and os.path.exists(path):
                            os.system("yes|rm -r "+str(path))
                        if create:
                            os.makedirs(path, exist_ok=True)

                    name = mystring.string("ITR>{0}_URL>{1}_STR>{2}\n".format(
                        repo_itr, repo_clone_url, search_string
                    ))
                    repo_dir = "repo_" + str(name.tobase64())
                    results_dir = "results_" + str(name.tobase64())

                    self.repair(repo_dir, create=False, delete_paths = self.delete_paths)
                    self.repair(results_dir)

                    sqlite_db_file = os.path.join(results_dir, "git_to_net.sqlite")

                    git2.clone_repository(repo_clone_url, repo_dir)  # Clones a non-bare repository

                    self.mine_repo(repo_dir, sqlite_db_file)

                    if os.stat(sqlite_db_file).st_size > 100_000_000:
                        with foldentre(new_path=results_dir):
                            raw_db_file = sqlite_db_file.replace(results_dir, '')
                            splittr.hash(raw_db_file)
                            splittr.split(raw_db_file, 50_000_000)
                            splittr.template(raw_db_file+".py")

                    appr(string(name.replace(',',';').replace('_',',').strip()))
                    fin_queue.put(repo_clone_url)

                return process

            if len(self.repos) > 0:
                for repo_itr, repo in enumerate(self.repos):
                    self.processor += process_prep(repo_itr, repo.ghub_url.furl, self.query_string, self.appr, self.processed_paths)
                    self.current_repo_itr = repo_itr
            else:
                print("No Repos Found")

        @property
        def repos(self):
            if self.tracking_repos is None:
                self.tracking_repos = []
                if os.path.exists(self.localfilename):
                    with open(self.localfilename, "r") as reader:
                        for line in reader:
                            ProjectItr, ProjectURL, ProjectScanned = line.split(",")
                            if ProjectScanned == "false":
                                self.tracking_repos.append(ProjectURL)
                else:
                    clean_url = lambda url:url.replace(".git", "").replace("https://", "").replace("http://","").replace("github.com/","")
                    if self.start_num == -1 and self.end_num == -1:
                        self.tracking_repos = [ghub(clean_url(x.clone_url), self.token, usewget=self.usewget) for x in self.g.search_repositories(query=self.query_string)]
                    else:
                        self.tracking_repos = []
                        start_page = 1 if self.start_num == -1 else self.start_num
                        for x_page, x in enumerate(self.g.search_repositories(query=self.query_string, page=start_page)):
                            if (self.end_num == -1 or self.end_num <= x_page):
                                self.tracking_repos += [
                                    ghub(clean_url(x.clone_url), self.token, usewget=self.usewget)
                                ]

                            if (
                                self.end_num != -1
                                and
                                self.end_num > x_page
                                and
                                self.end_num == x_page - 1
                            ):
                                break
            return self.tracking_repos

        @property
        def complete(self):
            return self.total_repo_len == self.current_repo_itr and self.processor.complete

        @property
        def localfilename(self):
            if self.tracking_name is None:
                self.tracking_name = mystring.string("query_progress_{0}.csv".format(
                    mystring.string("{0}".format(self.query_string)).tobase64())
                )
            return self.tracking_name

        @property
        def timing(self):
            self.api_watch.timing

            extra_rate_limiting = self.g.get_rate_limit()
            if hasattr(extra_rate_limiting, "search"):
                search_limits = getattr(extra_rate_limiting, "search")
                if search_limits.remaining < 2:
                    print("Waiting until: {0}".format(search_limits.reset))
                    pause.until(search_limits.reset)
except:pass

try:
    from gitlab import Gitlab
    class glab(mem):
        #https://python-gitlab.readthedocs.io/en/stable/index.html#installation
        #https://python-gitlab.readthedocs.io/en/stable/api-usage.html
        def __init__(self,repo,access_token=None,oauth=None,branch='master',wraplambda=lambda foil:False):
            super().__init__(wraplambda)
            if access_token:
                self.gl = Gitlab(private_token=access_token)
            elif oauth:
                self.gl = Gitlab(oauth_token=oauth)
            else:
                self.gl = Gitlab()

            self.gl.auth()

            if "/" in repo:
                self.reponame = repo.strip("/")[-1]
            else:
                self.reponame = repo

            self.project = None
            self.branch = branch

            for project in self.gl.projects.list(owned=True,get_all=True):
                if project.name == self.reponame:
                    self.project = project

        def files(self):
            #https://python-gitlab.readthedocs.io/en/stable/gl_objects/projects.html?highlight=files#project-files
            #https://stackoverflow.com/questions/60243129/how-can-i-extract-contents-from-a-file-stored-in-gitlab-repos

            files = [
                x['path'] for x in 
                self.project.repository_tree(ref=self.branch,all=True,recursive=True)
            ]

            return files

        def url(self):
            return "https://gitlab.com/" + str(self.repo)

        def login(self):
            return
        
        def logout(self):
            return
        
        def download(self, file_path=None,download_to=None):
            download_to = download_to or os.path.basename(file_path)
            #https://python-gitlab.readthedocs.io/en/stable/gl_objects/projects.html#id7

            if download_to is None:
                download_to = os.path.join(os.curdir,file_path.split("/")[-1])

            if file_path and isinstance(file_path,str):
                with open(download_to,"wb+") as writer:
                    self.project.files.raw(file_path=file_path, ref=self.branch,streamed=True,action=writer.write)
            return download_to
        
        def upload(self, file_path=None,path_in_repo=None):
            #https://python-gitlab.readthedocs.io/en/stable/gl_objects/projects.html
            with open(file_path, 'r') as my_file:
                file_content = my_file.read()

            if path_in_repo in self: #Update
                f = self.project.files.get(file_path=path_in_repo, ref=self.branch)
                f.content = file_content
                f.save(branch=self.branch, commit_message='Auto Update File')
            else: #Create https://stackoverflow.com/questions/52338343/problem-to-upload-a-file-to-gitlab-using-python-gitlab
                f = self.project.files.create({
                    'file_path': path_in_repo,
                    'branch': self.branch,
                    'content': file_content,
                    'author_email': self.gl.user.email,
                    'author_name': self.gl.user.name,
                    #'encoding': 'utf-8',
                    'commit_message': 'Auto Create File'
                })
        
        def delete_file(self,path_in_repo=None):
            if path_in_repo in self.files():
                self.project.files.delete(
                    file_path=path_in_repo,
                    branch=self.branch,
                    commit_message="Auto Delete From Repo"
                )
except: pass

try:
    from zipfile import ZipFile
    import ruamel.std.zipfile as zipfileextra
    class zyp(mem):
        #https://docs.python.org/3/library/zipfile.html
        def __init__(self,zyp_file,wraplambda=lambda foil:False):
            super().__init__(wraplambda)
            self.location = zyp_file

        def files(self):
            if not os.path.exists(self.location):
                return []
            return ZipFile(self.location).namelist()

        def login(self):
            return
        
        def logout(self):
            return
        
        def download(self, file_path=None,download_to=None):
            if not os.path.exists(self.location):
                print("Zip File Does Not Exist")
                return

            download_to = download_to or os.path.basename(file_path)
            with ZipFile(self.location) as z:
                print(download_to)
                with open(download_to, 'wb') as f:
                    f.write(z.read(file_path))
            return download_to
        
        def upload(self, file_path=None,path_in_repo=None):
            editing_mode = 'a' if os.path.exists(self.location) else 'w'

            if path_in_repo in self.files():
                del self[path_in_repo]

            with ZipFile(self.location,editing_mode) as myzip:
                myzip.write(file_path,path_in_repo)
            return True
        
        def delete_file(self,path_in_repo=None):
            if not os.path.exists(self.location):
                return

            zipfileextra.delete_from_zip_file(self.location, file_names=[path_in_repo])
            return True
except:pass

try:
    import tarfile
    class tar(mem):
        #https://docs.python.org/3/library/tarfile.html
        def __init__(self,file_path,wraplambda=lambda foil:False):
            super().__init__(wraplambda)
            self.location = file_path

        def files(self):
            #https://www.tutorialspoint.com/How-to-create-a-tar-file-using-Python
            files = []
            if os.path.exists(self.location):
                with tarfile.open(self.location, 'r') as tar:
                    files = tar.getnames()
            return files

        def login(self):
            return
        
        def logout(self):
            return
        
        def download(self, file_path=None,download_to=None,try_anyway=False):
            if not os.path.exists(self.location):
                print("Tar File Does Not Exist")
                return
            if file_path not in self.files() and not try_anyway:
                print("File Does Not Exist within tar")
                return
            
            try:
                found_file_name = file_path
                if try_anyway:
                    refound = False
                    cur_files = self.files()
                    for test_file_name in [file_path, os.path.basename(file_path)]:
                        if test_file_name in cur_files and not refound:
                            found_file_name = test_file_name
                            refound = True
                            break
                    if not refound:
                        for cur_file in cur_files:
                            if not refound:
                                for test_file_name in [file_path, os.path.basename(file_path)]:
                                    if cur_file.endswith(test_file_name) and not refound:
                                        found_file_name = test_file_name
                                        refound = True
                                        break

                with tarfile.open(self.location, 'r') as tar:
                    #https://stackoverflow.com/questions/20434912/is-it-possible-to-extract-single-file-from-tar-bundle-in-python
                    tar.extract(found_file_name, path=os.path.dirname(download_to))
            except Exception as e:
                print(e)


            return download_to
        
        def upload(self, file_path=None,path_in_repo=None):
            #https://stackoverflow.com/questions/2239655/how-can-files-be-added-to-a-tarfile-with-python-without-adding-the-directory-hi
            with tarfile.open(self.location, 'a' if os.path.exists(self.location) else 'w') as tar:
                tar.add(file_path, arcname=path_in_repo)

            return True
        
        def delete_file(self,path_in_repo=None):
            if not os.path.exists(self.location) or path_in_repo not in self.files():
                return
            backup_location = self.location + ".backup"

            with tarfile.open(self.location, 'r') as original:
                with tarfile.open(backup_location, 'w') as modified:
                    for info in original.getmembers():
                        if info.name == path_in_repo:
                            continue
                        extracted = original.extractfile(info)
                        if not extracted:
                            continue
                        modified.addfile(info, extracted)
                        os.remove(extracted.name)

            os.remove(self.location)
            os.rename(backup_location, self.location)
            
            return True
except: pass

try:
    import docker, tarfile
    from ephfile import ephfile
    class dock(mem):
        #https://docs.python.org/3/library/tarfile.html
        def __init__(self,container:docker.models.containers.Container,working_dir:str="/tmp",wraplambda=lambda foil:False):
            super().__init__(wraplambda)
            self.container = container
            self.working_dir = working_dir

        def files(self):
            #https://www.tutorialspoint.com/How-to-create-a-tar-file-using-Python
            exit_code=-1;files = []
            try:
                file_logs = self.container.exec_run(
                    cmd = "sh -c 'ls -altr {0}'".format(self.working_dir),
                    privileged=True,
                    workdir=self.working_dir,
                    stderr=True, stdout=True
                )
                for log_itr,file_log in enumerate(file_logs):
                    if log_itr == 0:
                        try:
                            exit_code = int(file_log.strip())
                        except:pass
                    else:
                        try:
                            file_log_line = str(file_log.decode("utf-8")).strip()
                            for subline in file_log_line.split("\n"):
                                if not subline.startswith("total"):
                                    subline = str(subline.split(" ")[-1]).strip()
                                    if subline not in ["../","..","./",".",""]:
                                        files += [subline]
                        except Exception as k:
                            print("Error  decoding {1} @ line {0}".format(str(log_itr), str(file_log)))
            except Exception as e:
                print(e)
            return files

        def _check_for_file(self, path):
            try:
                bits, stat = self.container.get_archive(path)
                return True
            except docker.errors.NotFound:
                return False
            except Exception as e:
                print(e)
                return False

        def login(self):
            return
        
        def logout(self):
            return
        
        def download(self, file_path=None, download_to=None):
            if file_path.replace(self.working_dir+"/", "") not in self.files():
                print("File {0} Does Not Exist within".format(file_path.replace(self.working_dir+"/", "")))
                return

            with ephfile(suffix=".tar",create=False) as temp_tar:
                with open(temp_tar(), "wb") as f:
                    bits, stat = self.container.get_archive(os.path.join(self.working_dir, file_path))
                    for chunk in bits:
                        f.write(chunk)

                with tar(temp_tar()) as tarfile:
                    try:
                        if file_path in tarfile.files():
                            dir_name = os.path.dirname(download_to).strip()
                            if dir_name != '':
                                os.makedirs(dir_name, exist_ok=True)

                            tarfile.download(file_path, download_to, try_anyway=True)
                    except Exception as e:
                        print(e)

            return download_to
        
        def upload(self, file_path=None,path_in_repo=None):
            with ephfile(suffix=".tar",create=False) as temp_tar:

                with tar(temp_tar()) as tar_file:
                    tar_file.upload(file_path, file_path)

                with open(temp_tar(), "rb") as in_file:
                    tar_file_bytes = in_file.read()

                self.container.put_archive(
                    self.working_dir,
                    tar_file_bytes
                )

            return True
        
        def delete_file(self,path_in_repo=None):
            try:
                del_logs = self.container.exec_run(
                    cmd = "sh -c 'yes|rm -r {0}'".format(path_in_repo),
                    privileged=True,
                    workdir=self.working_dir,
                    stderr=True, stdout=True
                )
                return True
            except Exception as e:
                print(e)
                return True
except: pass

try:
    from openpyxl import load_workbook
    class xcyl(mem):
        #https://python-gitlab.readthedocs.io/en/stable/index.html#installation
        #https://python-gitlab.readthedocs.io/en/stable/api-usage.html
        def __init__(self,path):
            self.path = path

        def files(self):
            return load_workbook(self.path, read_only=True, keep_links=True).sheetnames

        def login(self):
            return
        
        def logout(self):
            return

        def load(self, sheet=None):
            if file_path not in self.files():
                return None

            return pd.read_excel(self.path, sheet_name=file_path, engine="openpyxl")

        def download(self, file_path=None, download_to=None):
            if file_path not in self.files():
                return None

            data = pd.read_excel(self.path, sheet_name=file_path, engine="openpyxl")

            if download_to:
                mystring.frame.of(data).write_to(download_to)

            return data
        
        def upload(self, file_path=None,path_in_repo=None):
            while path_in_repo in self.files():
                path_in_repo += "_"

            data = mystring.frame.of(file_path)
            
            with pd.ExcelWriter(self.path, engine="xlsxwriter") as writer:
                data.to_excel(writer, sheet_name=path_in_repo, startrow=1, header=True)

            return True
        
        def delete_file(self,path_in_repo=None):
            if path_in_repo not in self.files():
                return False
            
            current_date = {}
            for sheet in self.files():
                if sheet != path_in_repo:
                    current_date[sheet] = pd.read_excel(self.path, sheet_name=sheet, engine="openpyxl")
            
            os.remove(self.path)

            with pd.ExcelWriter(self.path, engine="xlsxwriter") as writer:
                for sheet in current_date:
                    current_date[sheet].to_excel(writer, sheet_name=sheet, startrow=1, header=True)

            return True
except: pass

try:
    import pandas as pd
    import mystring as mys
    import sqlite3
    class sqlite(mem):
        #https://python-gitlab.readthedocs.io/en/stable/index.html#installation
        #https://python-gitlab.readthedocs.io/en/stable/api-usage.html
        def __init__(self,path):
            self.path = path

        def __crawl(self):
            class __internal_crawl(object):
                def __init__(self, file=None):
                    self.file = file
                    self.connection = None

                def __enter__(self):
                    self.connection = sqlite3.connect(self.file)
                    return self

                def __call__(self, query:str):
                    cursor = self.connection.cursor()
                    cursor.execute(query)
                    output = cursor.fetchall()
                    cursor = None
                    return output

                def __exit__(self, a,b,c):
                    if self.connection:
                        self.connection.close()
            return __internal_crawl(self.path)

        def files(self):
            tables = []
            with self.__crawl() as db:
                tables = [str(x[0]).strip() for x in db("SELECT name FROM sqlite_master WHERE type='table';")]
            return tables

        def login(self):
            return
        
        def logout(self):
            return
        
        def download(self, file_path=None,download_to=None):
            if file_path not in self.files():
                return None

            data = None
            with self.__crawl() as db:
                data = mys.frame(pd.read_sql_query("""SELECT * FROM "{0}";""".format(file_path), con=db.connection))
            
            if download_to:
                from pathlib import Path
                ext = Path(download_to).suffix

                if ext == ".pkl":
                    data.to_pickle(download_to)
                elif ext == ".csv":
                    data.to_csv(download_to)
                elif ext == ".excel":
                    data.to_excel(download_to)

            return data
        
        def upload(self, file_path=None,path_in_repo=None):
            if file_path is None:
                return False

            if isinstance(file_path, str):
                from pathlib import Path
                ext = Path(path_in_repo).suffix

                if not os.path.exists(file_path):
                    return False
                if ext == ".pkl":
                    file_path = pd.read_pickle(file_path)
                elif ext == ".csv":
                    file_path = pd.read_csv(file_path)
                elif ext == ".excel":
                    default_sheet_name = 0

                    if ":" in file_path:
                        file_path, default_sheet_name = file_path.split(":")

                    file_path = pd.read_excel(file_path, sheet_name = default_sheet_name)
                else:
                    return False

            if path_in_repo is None:
                path_in_repo = os.path.basename(file_path)

            current_file_names = list(self.files())
            while path_in_repo in current_file_names:
                path_in_repo += "_"

            with self.__crawl() as db:
                file_path.to_sql(self.path, db.connection, if_exists='replace')

            return True
        
        def delete_file(self,path_in_repo=None):
            if path_in_repo is None or path_in_repo not in list(self.files()):
                return True

            with self.__crawl() as db:
                db.connection.cursor().execute("""DROP table IF EXISTS "{0}";""".format(path_in_repo))

            return True
except: pass

try:
    from ephfile import ephfile
    import pydbhub.dbhub as dbhub

    class dbhub_repo(mem): #https://github.dev/franceme/xcyl
        #https://python-gitlab.readthedocs.io/en/stable/index.html#installation
        #https://python-gitlab.readthedocs.io/en/stable/api-usage.html
        def __init__(self, repo:str, access_token: str, owner: str, table_name: str=None):
            self.repo = repo
            self.access_token = access_token
            self.owner = owner
            self.config = "config.ini"
            self.table_name = table_name

        @property
        def dbhub(self):
            class login_prep(object):
                def __init__(self, api_key, db_owner, db_name):
                    self.api_key = api_key
                    self.db_owner = db_owner
                    self.db_name = db_name
                    self.config_file = "config.ini"
                def __enter__(self):
                    with open(self.config, "w+") as config:
                        config.write("""[dbhub]
	api_key = {0}
	db_owner = {self.owner}
	db_name = {self.repo}
	""".format(
        self.api_key,
        self.db_owner,
        self.db_name
    ))
                    return dbhub.Dbhub(config_file=self.config)
                def __exit__(self, a=None,b=None,c=None):
                    os.remove(self.config)
                    pass
            return login_prep(self.access_token, self.owner, self.table_name)

        def query(self, string):
            output = None

            with self.dbhub as db:
                results, err = db.Query(
                    self.owner,
                    self.repo,
                    string.replace(':table_name', self.table_name).replace(':table_name'.upper(), self.table_name)
                )
                if err is None:
                    output = results

            return output


        def files(self):
            output = []

            with self.dbhub as db:
                try:
                    # https://github.com/LeMoussel/pydbhub/blob/5fac7fa1b136ccdac09c58165ab399603c32b16f/examples/list_tables/main.py#L28
                    tables, err = db.Tables(self.owner, self.repo)
                    if err is None:
                        output = tables
                except:
                    pass

            return output

        def login(self):
            return
        
        def logout(self):
            return
        
        def download(self, file_path=None,download_to=None):
            data = self.query("SELECT * FROM {0}".format(file_path))

            if download_to:
                mystring.frame.of(data).write_to(download_to)

            return data
        
        def upload(self, file_path=None, path_in_repo=None):
            if file_path == None:
                return
            while path_in_repo in self.files():
                path_in_repo += "_"

            if isinstance(file_path, str):
                new_table = mystring.frame.of(file_path)

            with ephfile(suffix='.sqlite') as eph:
                for table in self.files():
                    self.download(table, eph())
                new_table.write_to(eph(), sheet_name = path_in_repo)

            with self.dbhub as db:
                try:
                    db_contents = open(eph(), 'rb')
                    with db_contents:
                        # https://github.com/LeMoussel/pydbhub/blob/5fac7fa1b136ccdac09c58165ab399603c32b16f/examples/upload/main.py#L51
                        res, err = db.Upload(db_name=eph(), db_bytes=db_contents,
                                            info=dbhub.UploadInformation(
                                                commitmsg=f"Uploading changes to {self.repo}",
                                                committimestamp=datetime.datetime.now(),
                                            ))
                        if err is not None:
                            print(f"[ERROR] {err}")
                            sys.exit(1)
                except Exception as e:
                    pass
            return
        
        def delete_file(self,path_in_repo=None):
            self.query("DROP TABLE IF EXISTS {0}".format(path_in_repo))
except: pass

def redundant(klass):
    """
    #Example:
    source = hugg.redundant([
        hugg.face(xxx, yyy),
        hugg.ghub(xxx, yyy),
    ])
    #One of the two provided should work, otherwise an exception is thrown
    """
    if not isinstance(klass,list):
        klass = [klass]

    for temp_klass in klass:
        try:
            temp_klass.files()
            return temp_klass
        except: pass
    
    raise Exception("No Mirrors are available")

def sync_two_repos(new_repo, old_repo,delay_sec=2, delete_old=False):
    import time
    delay = lambda:time.sleep(delay_sec)

    print("Updating/Creating New Files")
    for foil in new_repo.files():
        try:
            old_repo[foil] = new_repo[foil]
        except: pass
        delay()
        print(".",end='',flush=True)

    if delete_old:
        print("\nDeleting Old Files")
        for foil in old_repo.files():
            if foil not in new_repo.files():
                del old_repo[foil]
                delay()
            print(".",end='',flush=True)
    
    print("Completed")

def convert_size_to_bytes(size_str):
    try:
        #https://stackoverflow.com/questions/44307480/convert-size-notation-with-units-100kb-32mb-to-number-of-bytes-in-python
        """Convert human filesizes to bytes.

        Special cases:
        - singular units, e.g., "1 byte"
        - byte vs b
        - yottabytes, zetabytes, etc.
        - with & without spaces between & around units.
        - floats ("5.2 mb")

        To reverse this, see hurry.filesize or the Django filesizeformat template
        filter.

        :param size_str: A human-readable string representing a file size, e.g.,
        "22 megabytes".
        :return: The number of bytes represented by the string.
        """
        multipliers = {
            'kilobyte':  1024,
            'megabyte':  1024 ** 2,
            'gigabyte':  1024 ** 3,
            'terabyte':  1024 ** 4,
            'petabyte':  1024 ** 5,
            'exabyte':   1024 ** 6,
            'zetabyte':  1024 ** 7,
            'yottabyte': 1024 ** 8,
            'kb': 1024,
            'mb': 1024**2,
            'gb': 1024**3,
            'tb': 1024**4,
            'pb': 1024**5,
            'eb': 1024**6,
            'zb': 1024**7,
            'yb': 1024**8,
        }

        for suffix in multipliers:
            size_str = size_str.lower().strip().strip('s')
            if size_str.lower().endswith(suffix):
                return int(float(size_str[0:-len(suffix)]) * multipliers[suffix])
        else:
            if size_str.endswith('b'):
                size_str = size_str[0:-1]
            elif size_str.endswith('byte'):
                size_str = size_str[0:-4]
        return int(size_str)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info();fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        msg = ":> Hit an unexpected error {0} @ {1}:{2}".format(e, fname, exc_tb.tb_lineno)
        print(msg)
        return float("inf")

def available_types():
    storage_types = {}
    storage_types[":local"] = localdrive
    try:
        from huggingface_hub import HfApi;
        storage_types[":face"] = face
    except:pass
    try:
        from github import Github,Requester
        storage_types[":ghub"] = ghub
    except:pass
    try:
        from gitlab import Gitlab
        storage_types[":glab"] = glab
    except:pass
    try:
        import ruamel.std.zipfile as zipfileextra
        storage_types[".zip"] = zyp
    except:pass
    try:
        import tarfile
        storage_types[".tar"] = tar
    except:pass
    try:
        import docker, tarfile
        storage_types[":dock"] = dock
    except:pass
    try:
        from openpyxl import load_workbook
        storage_types[".xlsx"] = xcyl
    except:pass
    try:
        import pandas as pd
        import mystring as mys
        import sqlite3
        storage_types[".sqlite"] = sqlite
    except:pass
    try:
        import pydbhub.dbhub as dbhub
        storage_types[":dbhub"] = dbhub_repo
    except:pass
    return storage_typs

class eph_mgr(object):
    def __init__(self,repo:mem=None, load_all=False, *files):
        self.repo = repo
        self.files = {}
        self.load_all = load_all
        for file in files:
            self.files[file] = None

    def __getitem__(self, key):
        output = None
        if key in self.files:
            output = self.files[key]
        return output

    def __enter__(self):
        total_files = list(self.repo.files())
        if self.load_all:
            import re
            lyst = []
            matching_prefix = 'r:'

            for files_key in list(self.files.keys()):
                if files_key.startswith(matching_prefix):
                    files_key = files_key.replace(matching_prefix,'')
                    lyst += [
                        x for x in total_files if re.match(files_key, x) is not None
                    ]
                else:
                    lyst += [
                        files_key
                    ]
        else:
            lyst = total_files

        for file_name in lyst:
            if file_name.endswith(".py"):
                self.files[
                    os.path.basename(file_name).replace(".py","")
                ] = self.repo.impor(file_name, delete=True)
            else:
                self.files[os.path.basename(file_name)] = file_name

        return self

    def __exists(self, file):
        return file in list(self.files.keys())

    def __exit__(self, a=None,b=None,c=None):
        get_folder = lambda path:os.path.abspath(path).replace(os.path.basename(path),"")
        for foil in self.files:
            try:
                if os.path.exists(foil):
                    os.remove(foil)

                folder = get_folder(foil)
                if len(os.listdir(folder)) == 0:
                    os.rmdir(folder)
            except:pass

    #^=	.__ixor__(self, other)
    def __ixor__(self, file_name:str):
        if self.__exists(file_name):
            self.download(file_name, file_name)
            return fput(file_name)
        return None
