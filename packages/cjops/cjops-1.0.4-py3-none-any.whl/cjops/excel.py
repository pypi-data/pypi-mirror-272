from loguru import logger
import pendulum
from pathlib import Path
from openpyxl import Workbook
from typing import List, Dict, Any, Tuple

def write_list_tuple_excel(datas: List[Tuple[Any]]|List[List[Tuple[Any]]], headers: List[str] = None, filename: str = None, sheet_name: str|List|None=None) -> None:
    """
    将一个或者多个列表元组写入到excel文件中
    datas:
        eg1. [('argo', 'elk'),('argo2', 'elk2')]
        eg2. [[('argo', 'elk'),('argo2', 'elk2')], [(...), (...)]]
    headers: 表头列表,默认None,不写表头; ['命名空间', '项目名称']
    filename: 写入excel中的文件名,默认为None
    sheet_name: 工作表名称,默认None;
        eg0: "hzUsa"
        eg1: 字符串传递: "杭州,美国"
        eg2: 列表传递: ["杭州","美国"]
    """
    # 判断文件是否已经存在,如果存在则删除
    if filename:
        if Path(filename).exists() or Path(f'{filename}.xlsx').exists():
            try:
                Path(filename).unlink()
                Path(f'{filename}.xlsx').unlink()
            except Exception:
                pass

    wb = Workbook()
    wb.encoding = 'utf-8'

    # 如果是多个list,则走下面的逻辑
    if isinstance(datas[0], list):
        for index, data in enumerate(datas, start=1):
            if sheet_name:
                try:
                    # 如果sheet_name是字符串
                    if isinstance(sheet_name, str):
                        ws = wb.create_sheet(title=sheet_name.split(',')[index-1])
                    if isinstance(sheet_name, list):
                        ws = wb.create_sheet(title=sheet_name[index-1])
                except Exception as e:
                    logger.error(f'sheet_name获取失败,采用默认sheet_name,错误信息: {e}')
                    ws = wb.create_sheet(title=f"Sheet{index}")
            else:
                ws = wb.create_sheet(title=f"Sheet{index}")

            if headers:
                ws.append(headers)
            for item in data:
                ws.append(item)
            # 设置单元格宽度
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # 删除默认的Sheet
        wb.remove(wb['Sheet'])

    # 如果是单个list,则走下面的逻辑
    if isinstance(datas[0], tuple):
        if sheet_name:
            try:
                # 如果sheet_name是字符串
                if isinstance(sheet_name, str):
                    ws = wb.create_sheet(title=str(sheet_name))
                if isinstance(sheet_name, list):
                    ws = wb.create_sheet(title=sheet_name[0])
                # 删除默认的Sheet
                wb.remove(wb['Sheet'])
            except Exception as e:
                logger.error(f'sheet_name获取失败,采用默认sheet_name,错误信息: {e}')
                ws = wb.active
        else:
            ws = wb.active

        if headers:
            ws.append(headers)
        for item in datas:
            ws.append(item)
        # 设置单元格宽度
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # 自动添加后缀xlsx
    if filename and not str(filename).endswith('.xlsx'):
        filename = filename +  '.xlsx'

    if not filename:
        filename = f'write-{pendulum.now().to_datetime_string()}.xlsx'

    wb.save(filename)
    logger.success(f'写入文件成功,文件名: {filename}')

def write_list_dict_excel(datas: List[Dict[str, Any]]|List[List[Dict[str, Any]]], headers: Dict[str, str]=None, filename: str=None, sheet_name: str|List|None=None):
    """
    将一个或者多个字典列表写入excel中
    datas: eg1. [{'ns': 'argo', 'tp': 'elk'}, {'ns': 'argo2', 'tp': 'elk2'}]
           eg2. [[{'ns': 'argo', 'tp': 'elk'}, {'ns': 'argo2', 'tp': 'elk2'}], [{...}, {...}]]
    headers: 字典，每个键值对对应列名和标题；如果未提供，则根据第一个数据字典的键值自动生成; {'ns': '命名空间', 'tp': '类型'}
    filename: 响应头中的文件名,提供前端获取
    sheet_name: 工作表名称,默认None;
        eg0: "hzUsa"
        eg1: 字符串传递: "杭州,美国"
        eg2: 列表传递: ["杭州","美国"]
    """
    wb = Workbook()
    wb.encoding = 'utf-8'

    # 如果传递多个列表,则使用下面的创建逻辑
    if isinstance(datas[0], list):
        for index, data in enumerate(datas, start=1):
            if sheet_name:
                try:
                    # 如果sheet_name是字符串
                    if isinstance(sheet_name, str):
                        ws = wb.create_sheet(title=sheet_name.split(',')[index-1])
                    if isinstance(sheet_name, list):
                        ws = wb.create_sheet(title=sheet_name[index-1])
                except Exception as e:
                    logger.error(f'sheet_name获取失败,采用默认sheet_name,错误信息: {e}')
                    ws = wb.create_sheet(title=f"Sheet{index}")
            else:
                ws = wb.create_sheet(title=f"Sheet{index}")

            if headers and isinstance(headers, dict):
                header_values = [headers.get(key, key) for key in data[0].keys()]
                ws.append(header_values)
            else:
                ws.append(list(data[0].keys()))

            for item in data:
                ws.append(list(item.values()))

            # 设置单元格宽度
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # 删除默认的Sheet
        wb.remove(wb['Sheet'])

    # 如果是单个数据list,走下面的逻辑
    if isinstance(datas[0], dict):
        if sheet_name:
            try:
                # 如果sheet_name是字符串
                if isinstance(sheet_name, str):
                    ws = wb.create_sheet(title=str(sheet_name))
                if isinstance(sheet_name, list):
                    ws = wb.create_sheet(title=sheet_name[0])
                # 删除默认的Sheet
                wb.remove(wb['Sheet'])
            except Exception as e:
                logger.error(f'sheet_name获取失败,采用默认sheet_name,错误信息: {e}')
                ws = wb.active
        else:
            ws = wb.active

        if headers and isinstance(headers, dict):
            header_values = [headers.get(key, key) for key in datas[0].keys()]
            ws.append(header_values)
        else:
            ws.append(list(datas[0].keys()))

        for item in datas:
            ws.append(list(item.values()))

        # 设置单元格宽度
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # 自动添加后缀xlsx
    if filename and not str(filename).endswith('.xlsx'):
        filename = filename +  '.xlsx'

    if not filename:
        filename = f'write-{pendulum.now().to_datetime_string()}.xlsx'

    wb.save(filename)
    logger.success(f'写入文件成功,文件名: {filename}')
